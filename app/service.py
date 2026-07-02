"""
service.py — оркестратор генерации документов.

Единый конвейер (Спринт 2):
    файлы → sources.read_shipment → canonical.Shipment → заполнители/сборщики.

INV(RU), PL и «Весь шаблон» — это ОДНО заполнение OptiAuto-workbook (листы
INV+PL) через общий движок; INV(RU)/PL получаются извлечением нужного листа.
Транзит и CMR строятся из canonical через si_bridge (работают даже когда весь
шаблон не заполнен). Шаблон и режим шапки выбираются по ОТПРАВИТЕЛЮ, а не ручным
полем «тип шапки».

Каждый тип вывода обёрнут в try/except: ошибка одного документа не ломает
остальные — пользователь получает то, что удалось, плюс понятные сообщения.
"""
from __future__ import annotations

import datetime
import re
import sys
import traceback
from pathlib import Path

from .config import ENGINES_DIR

# Легаси-движки используют импорты верхнего уровня — папку engines кладём в path.
if str(ENGINES_DIR) not in sys.path:
    sys.path.insert(0, str(ENGINES_DIR))

import openpyxl  # noqa: E402
import generate_si_documents as si  # noqa: E402
import hssc_txt_v3 as hssc_engine  # noqa: E402
import parties as parties_db  # noqa: E402

from .engines.sources import read_shipment  # noqa: E402
from .engines.fillers import fill_optiauto_whole  # noqa: E402
from .engines.si_bridge import build_si_inputs  # noqa: E402
from .engines.template_select import select_template  # noqa: E402


def parties_for_ui() -> dict:
    """Реестр отправителей/получателей для выпадающих списков в UI.
    Каждый элемент: {key, name, block} — block это готовый текст для textarea.
    """
    def pack(table):
        items = []
        for key, p in table.items():
            block = "\n".join([p["name"], *p.get("lines", [])])
            items.append({"key": key, "name": p["name"], "block": block})
        return items
    return {
        "senders": pack(parties_db.SENDERS),
        "receivers": pack(parties_db.RECEIVERS),
    }


# Доступные типы вывода (ключ → подпись для UI). Сверху — три основных документа.
OUTPUT_TYPES = {
    "inv_ru":           "INV (RU) — инвойс с русскими наименованиями",
    "pl":               "PL — упаковочный лист",
    "transit_document": "Invoice Transit (INVOICE_TRANSIT)",
    "txt_hssc":         "txt HSSC codes (таможня Дубая)",
    "cmr_document":     "CMR document (таблица для накладной)",
    "whole_template":   "Весь шаблон (заполненный INV + PL)",
}


def _safe(name: str) -> str:
    """Безопасное имя файла."""
    return re.sub(r"[^A-Za-z0-9_.\-]+", "_", str(name)).strip("_") or "file"


def _lines(text: str | None) -> list[str]:
    """Многострочное поле формы → список непустых строк."""
    if not text:
        return []
    return [ln.strip() for ln in text.replace("\r\n", "\n").split("\n") if ln.strip()]


class Result:
    def __init__(self):
        self.created: list[Path] = []   # успешно созданные файлы
        self.messages: list[str] = []   # предупреждения / ошибки для пользователя

    def add(self, path: Path):
        self.created.append(path)

    def warn(self, msg: str):
        self.messages.append(msg)


def _extract_sheet(filled_path: Path, keep_sheet: str, out_path: Path) -> None:
    """Скопировать заполненный workbook, оставив только один лист (INV или PL)."""
    wb = openpyxl.load_workbook(filled_path)
    for name in list(wb.sheetnames):
        if name != keep_sheet:
            del wb[name]
    wb.save(out_path)
    wb.close()


# --------------------------------------------------------------------------- #
# Точка входа
# --------------------------------------------------------------------------- #

def generate_documents(
    *,
    source_path: str | Path,
    selections: list[str],
    sender_text: str | None,
    receiver_text: str | None,
    out_dir: str | Path,
    invoice_no: str | None = None,
    date: str | None = None,
    sender_key: str | None = None,
    receiver_key: str | None = None,
) -> Result:
    src = Path(source_path)
    out_dir = Path(out_dir)
    out_dir.mkdir(parents=True, exist_ok=True)
    stem = _safe(src.stem)
    sender_lines = _lines(sender_text)
    receiver_lines = _lines(receiver_text)
    inv_label = _safe(invoice_no) if invoice_no else stem
    res = Result()

    # Нормализуем к .xlsx (.xls/.xlsb → LibreOffice). openpyxl-выводам нужен .xlsx.
    try:
        src_xlsx = si.to_xlsx_if_needed(src)
    except Exception:
        src_xlsx = src

    # ---- Распознаём исходник в canonical (один раз, лениво) ----
    _cache: dict = {}

    def shipment():
        if "sh" not in _cache:
            sh = read_shipment([src_xlsx])
            sh.requisites = sh.requisites.merged(
                invoice_no=invoice_no, date=date,
                sender_lines=sender_lines, receiver_lines=receiver_lines,
                sender_key=sender_key, receiver_key=receiver_key,
            )
            _cache["sh"] = sh
        return _cache["sh"]

    def si_data():
        """agg/details/profile/places/inv_no из canonical (кэш для transit+cmr)."""
        if "si" not in _cache:
            sh = shipment()
            inv_df, pl_df, details, profile = build_si_inputs(sh)
            if invoice_no:
                details["invoice_no"] = invoice_no
            if date:
                details["date"] = date
            si.PROFILES[profile]["sender"] = sh.requisites.sender_lines or []
            si.PROFILES[profile]["default_consignee"] = sh.requisites.receiver_lines or []
            details["consignee"] = sh.requisites.receiver_lines or details.get("consignee") or []
            enriched = si.enrich(inv_df)
            agg = si.aggregate(enriched)
            total_places = len(pl_df)
            agg = si.distribute_places(agg, total_places, enriched)
            inv_no = details.get("invoice_no") or src.stem
            _cache["si"] = (agg, details, profile, total_places, inv_no)
        return _cache["si"]

    # ---- Заполнение шаблона (INV(RU) / PL / Весь шаблон — одно заполнение) ----
    fill_selected = [k for k in ("inv_ru", "pl", "whole_template") if k in selections]
    if fill_selected:
        try:
            sh = shipment()
            choice = select_template(sender_key)
            whole_path = out_dir / f"WHOLE_{inv_label}.xlsx"
            rep = fill_optiauto_whole(
                sh, choice.whole_template, whole_path,
                invoice_no=invoice_no, date=date, remove_logo=choice.remove_logo)
            for note in choice.notes + rep.get("notes", []):
                res.warn(note)

            if "inv_ru" in selections:
                p = out_dir / f"INV_RU_{inv_label}.xlsx"
                _extract_sheet(whole_path, "INV", p)
                res.add(p)
            if "pl" in selections:
                p = out_dir / f"PL_{inv_label}.xlsx"
                _extract_sheet(whole_path, "PL", p)
                res.add(p)
            if "whole_template" in selections:
                res.add(whole_path)
            else:
                whole_path.unlink(missing_ok=True)   # был лишь промежуточным
        except Exception as e:
            res.warn(f"Заполнение шаблона: не удалось — {e}")
            traceback.print_exc()

    # ---- Транзитный документ ----
    if "transit_document" in selections:
        try:
            agg, details, profile, total_places, inv_no = si_data()
            p = out_dir / f"INVOICE_TRANSIT_{_safe(inv_no)}.xlsx"
            si.build_invoice_transit(
                agg, details, profile, p,
                blank=False, display_invoice_no=inv_no)
            res.add(p)
        except Exception as e:
            res.warn(f"Транзитный документ: не удалось — {e}")
            traceback.print_exc()

    # ---- CMR ----
    if "cmr_document" in selections:
        try:
            agg, details, profile, total_places, inv_no = si_data()
            p = out_dir / f"CMR_{_safe(inv_no)}.xlsx"
            si.build_cmr_table(agg, details, profile, total_places, p, blank=False)
            res.add(p)
        except Exception as e:
            res.warn(f"CMR: не удалось — {e}")
            traceback.print_exc()

    # ---- TXT HSSC ----
    if "txt_hssc" in selections:
        try:
            if src_xlsx.suffix.lower() != ".xlsx":
                raise RuntimeError(
                    "нужен .xlsx (для .xls/.xlsb установите LibreOffice)")
            wb = openpyxl.load_workbook(src_xlsx, data_only=True)
            iso2_to_name = {iso2: full
                            for full, iso2 in hssc_engine.COUNTRY_NAME_TO_ISO2.items()}
            sheet = next((c for c in ("HS CODE SUM", "HSSC") if c in wb.sheetnames), None)
            if sheet:
                agg_rows = hssc_engine.aggregate_hs_code_sum(wb[sheet], iso2_to_name)
            else:
                inv_name = next((c for c in ("INV", "INVOICE") if c in wb.sheetnames), None)
                if not inv_name:
                    raise RuntimeError("нет листа HS CODE SUM / HSSC / INV")
                agg_rows = hssc_engine.aggregate_inv(wb[inv_name])
            wb.close()
            if not agg_rows:
                raise RuntimeError("не найдено строк с HS-кодами")
            company = sender_lines[0] if sender_lines else "WENS LOGISTICS DWC-LLC"
            txt = hssc_engine.generate_txt(
                agg_rows,
                decl_no=stem,
                decl_date=datetime.date.today().strftime("%Y-%m-%d"),
                pages=1,
                company=company,
                supplier_code=1,
                incoterm="EXW",
                hs_units={},   # TODO Спринт 5: реальная база dt_hscodes (CD_BLANK_NEW.xlsx)
            )
            p = out_dir / f"{stem}_HSSC.txt"
            p.write_bytes(txt.encode("utf-8"))
            res.add(p)
        except Exception as e:
            res.warn(f"TXT HSSC: не удалось — {e}")

    if not res.created and not res.messages:
        res.warn("Не выбран ни один тип документа.")
    return res
