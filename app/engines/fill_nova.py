"""
fill_nova.py — runner for the proforma-enriched fill clients (LLCNova, AUTO-EURO).

⚠️  ЗАВИСИМОСТЬ: этот модуль вызывает fill_optiauto() с параметрами
    `inv_enrichment` и `pl_dedup_field`. В перенесённой версии `fill_optiauto.py`
    этих параметров НЕТ (она старее). Для работы fill_nova нужно заменить
    `fill_optiauto.py` актуальной версией. См. ROADMAP — Спринт 2.

These clients reuse the OptiAuto INV+PL template (logos removed, addresses
replaced — see NOVA_template.xlsx / AUTOEURO_template.xlsx) but differ from
OptiAuto/Carbotec in two ways that this runner handles:

  1. The commercial invoice carries ENGLISH descriptions only. The Russian
     description and the per-unit weight (and, for AUTO-EURO, the HS code —
     its invoice has NO HS column at all) are looked up from a separate
     PROFORMA workbook, keyed by article (invoice ITEM CODE ↔ proforma Артикул).

  2. The packing list lists one row per (item × case): only the first row of
     each case carries the box W/H/L/CBM/weight, the remaining rows of that
     case are zero. The boxes are collapsed to one row per case via
     fill_optiauto's pl_dedup_field='pkg_number'.

────────────────────────────────────────────────────────────────────────────
CRITICAL — READ THE PROFORMA IN FULL
────────────────────────────────────────────────────────────────────────────
The proforma must be read across EVERY row (2 .. ws.max_row). An earlier bug
capped the read at ~520 rows, which silently dropped RU descriptions / HS for
large proformas (e.g. проформа_1242 has 2030 rows / 2004 articles) and produced
English-fallback descriptions and blank HS codes. The small first proforma
(158 rows) fit under the cap by luck, which masked the bug. build_proforma_
enrichment() below always iterates to ws.max_row — never reintroduce a fixed
row cap.

The fill report's `blank_desc_filled` count is the canary: with a complete,
correctly-paired proforma it should be 0. A non-zero value means some invoice
articles were not found in the proforma (wrong/partial proforma, or an
article-format mismatch).
────────────────────────────────────────────────────────────────────────────

Profiles (company_profiles.py):
  llcnova  — ELITEGOODS TRADING LLC-FZ → LLC NOVA, EXW Dubai, USD.
             Invoice HAS its own HS column → include_hs_from_proforma=False.
  autoeuro — ProConstruct Limited → «AUTO-EURO» JSC, DAP, USD.
             Invoice has NO HS column → include_hs_from_proforma=True
             (HS taken from proforma ТНВЭД, truncated to 8 digits).

Source-sheet defaults (calibrated against the supplied files):
  invoice  : sheet 'Sheet_1', column header auto-detected (SR.NO / ITEM CODE /
             DESCRIPTION / SHIP QTY / PRICE / AMOUNT / COUNTRY [/ HS CODE]).
  proforma : sheet 'Sheet1', col 1 ТНВЭД, col 3 Артикул, col 4 Наименование
             рус., col 8 Unit Wt.
  packing  : sheet 'Sheet_1', column header auto-detected; CASE NO is the box
             id, Width/Height/Length on the row below the main header.
"""

from __future__ import annotations
from pathlib import Path
import argparse
import datetime as _dt

import openpyxl

from fill_optiauto import fill_optiauto


# --------------------------------------------------------------------------- #
# Proforma enrichment (FULL read)
# --------------------------------------------------------------------------- #

def build_proforma_enrichment(
    proforma_path: str | Path,
    *,
    sheet: str = "Sheet1",
    article_col: int = 3,      # Артикул
    ru_desc_col: int = 4,      # Наименование рус.
    unit_wt_col: int = 8,      # Unit Wt
    hs_col: int = 1,           # ТНВЭД (10-digit)
    hs_digits: int = 8,
    include_hs: bool = False,
) -> dict[str, dict]:
    """Return {article: {description, unit_weight[, hs_code]}} from a proforma.

    Reads the ENTIRE sheet (row 2 .. max_row) — see the module docstring on why
    a fixed row cap must never be reintroduced. hs_code is the ТНВЭД value
    truncated to `hs_digits` (only added when include_hs=True, i.e. when the
    invoice has no HS column of its own).
    """
    ws = openpyxl.load_workbook(proforma_path, data_only=True)[sheet]
    enrich: dict[str, dict] = {}
    for r in range(2, ws.max_row + 1):
        art = ws.cell(row=r, column=article_col).value
        if art in (None, ""):
            continue
        rec = {
            "description": ws.cell(row=r, column=ru_desc_col).value,
            "unit_weight": ws.cell(row=r, column=unit_wt_col).value,
        }
        if include_hs:
            tn = ws.cell(row=r, column=hs_col).value
            rec["hs_code"] = (
                str(tn).strip()[:hs_digits] if tn not in (None, "") else None)
        enrich[str(art).strip()] = rec
    return enrich


# --------------------------------------------------------------------------- #
# Date parsing helper
# --------------------------------------------------------------------------- #

def _parse_date(d):
    """Accept a datetime, or a string like '29 May 2026' / '2026-05-29' /
    '29.05.2026'."""
    if isinstance(d, _dt.datetime):
        return d
    if isinstance(d, _dt.date):
        return _dt.datetime(d.year, d.month, d.day)
    for fmt in ("%d %B %Y", "%Y-%m-%d", "%d.%m.%Y", "%d/%m/%Y"):
        try:
            return _dt.datetime.strptime(str(d).strip(), fmt)
        except ValueError:
            continue
    raise ValueError(f"Unrecognized date: {d!r}")


# --------------------------------------------------------------------------- #
# Public entry point
# --------------------------------------------------------------------------- #

def fill_nova(
    *,
    template_path: str | Path,
    source_inv_path: str | Path,
    source_pl_path: str | Path,
    proforma_path: str | Path,
    output_path: str | Path,
    invoice_no: str,
    invoice_date,
    currency: str = "USD",
    include_hs_from_proforma: bool = False,
    source_inv_sheet: str = "Sheet_1",
    source_pl_sheet: str = "Sheet_1",
    proforma_sheet: str = "Sheet1",
) -> dict:
    """Fill a proforma-enriched client (LLCNova / AUTO-EURO).

    The template must already carry the correct addresses / logo-removal for
    the client (NOVA_template.xlsx, AUTOEURO_template.xlsx). This runner sets
    the Date / Invoice No / Currency header cells, enriches the INV from the
    proforma (RU description + unit weight [+ HS]), and dedups the PL by CASE NO.

    Returns the fill_optiauto report dict; check report['blank_desc_filled'] —
    it should be 0 with a complete, correctly-paired proforma.
    """
    enrich = build_proforma_enrichment(
        proforma_path, sheet=proforma_sheet, include_hs=include_hs_from_proforma)

    report = fill_optiauto(
        template_path=template_path,
        source_inv_path=source_inv_path,
        source_pl_path=source_pl_path,
        output_path=output_path,
        source_inv_sheet=source_inv_sheet,
        source_pl_sheet=source_pl_sheet,
        inv_constants={"vat_rate": 0, "vat_amt": 0},
        pl_constants={"qty": 1},
        inv_header_updates={
            (14, 14): _parse_date(invoice_date),
            (15, 14): invoice_no,
            (18, 14): currency,
        },
        inv_enrichment=enrich,
        pl_dedup_field="pkg_number",
    )
    report["proforma_articles"] = len(enrich)
    return report


# --------------------------------------------------------------------------- #
# CLI
# --------------------------------------------------------------------------- #

def _main(argv=None):
    p = argparse.ArgumentParser(
        description="Fill a proforma-enriched client (LLCNova / AUTO-EURO).")
    p.add_argument("--template", required=True,
                   help="client template, e.g. NOVA_template.xlsx / "
                        "AUTOEURO_template.xlsx")
    p.add_argument("--invoice", required=True, help="source commercial invoice")
    p.add_argument("--packing", required=True, help="source packing list")
    p.add_argument("--proforma", required=True, help="source proforma")
    p.add_argument("--output", required=True, help="output .xlsx path")
    p.add_argument("--invoice-no", required=True)
    p.add_argument("--date", required=True,
                   help="invoice date, e.g. '29 May 2026' or 2026-05-29")
    p.add_argument("--currency", default="USD")
    p.add_argument("--hs-from-proforma", action="store_true",
                   help="invoice has no HS column → take HS from proforma "
                        "ТНВЭД (8-digit). Use for AUTO-EURO.")
    p.add_argument("--inv-sheet", default="Sheet_1")
    p.add_argument("--pl-sheet", default="Sheet_1")
    p.add_argument("--proforma-sheet", default="Sheet1")
    a = p.parse_args(argv)

    rep = fill_nova(
        template_path=a.template,
        source_inv_path=a.invoice,
        source_pl_path=a.packing,
        proforma_path=a.proforma,
        output_path=a.output,
        invoice_no=a.invoice_no,
        invoice_date=a.date,
        currency=a.currency,
        include_hs_from_proforma=a.hs_from_proforma,
        source_inv_sheet=a.inv_sheet,
        source_pl_sheet=a.pl_sheet,
        proforma_sheet=a.proforma_sheet,
    )
    print(f"Output:            {rep['output']}")
    print(f"INV rows:          {rep['inv_rows']}")
    print(f"PL boxes:          {rep['pl_rows']}")
    print(f"Proforma articles: {rep['proforma_articles']}")
    print(f"EN-fallback (RU):  {rep['blank_desc_filled']}  "
          f"({'OK' if rep['blank_desc_filled'] == 0 else 'CHECK proforma pairing/coverage'})")


if __name__ == "__main__":
    _main()
