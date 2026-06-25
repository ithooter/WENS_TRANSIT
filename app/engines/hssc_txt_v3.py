"""
WENS Logistic — TXT customs declaration generator (v3).

Final TXT format calibrated against reference TPL2605012.txt (29-May-2026).

Key format rules:
  IH header (15 fields):
    "IH","<decl_no>","<YY-MM-DD>","<pages>","1","<supplier_company>",
    "<supplier_code>","1","USD","<total>","<incoterm>","","","",""

  ID row (18 fields):
    "ID","<n>","<hs>","<qty> pcs AUTOSPARE PARTS","N",
    "<unit_label>","<unit_qty>","kg","<weight>",
    "","","<amount>","<iso2>","","","","",""

  Where for a given (HS, Country) aggregated row:
    unit_label, unit_qty depend on dt_hscodes lookup:
      kg / Square Meters / missing  -> label, weight    (field 6 = weight)
      u                              -> "u",  qty       (field 6 = qty)
      inv / na                       -> "?",  "?"

  Number formatting:
    weights and amounts use `.10g` (trim trailing zeros, no scientific):
      4737.000 -> "4737"        2194.20 -> "2194.2"
      71.900   -> "71.9"        2629.0  -> "2629"
      0.032    -> "0.032"

Aggregation:
  Rows are grouped by (HS code + Country ISO2). Within a pair, qty / amount /
  weight are summed; pairs are unique. Sort: (hs asc, iso2 asc).

Public entry points:
  load_hs_units(cd_blank_path)      -> {hs_code: unit_label}
  aggregate_inv(inv_ws)             -> [{hs, country_full, country_iso2,
                                         qty, amount, weight}]
  aggregate_hs_code_sum(ws)         -> same shape, reads pre-aggregated sheet
  write_hssc_sheet(wb, agg_rows)    -> fills HSSC in WENS template
  write_piv_hsc_sheet(wb, agg_rows) -> fills PIV HSC pivot
  generate_txt(...)                 -> returns TXT bytes
"""

from __future__ import annotations
import io
import math
from collections import OrderedDict
from pathlib import Path

import openpyxl

# --------------------------------------------------------------------------- #
# 1. HS unit lookup from dt_hscodes
# --------------------------------------------------------------------------- #

def load_hs_units(cd_blank_path: str | Path) -> dict[str, str]:
    """Return {hs_code_str: unit_desc_str} from CD_BLANK_NEW.xlsx dt_hscodes."""
    wb = openpyxl.load_workbook(cd_blank_path, data_only=True, read_only=True)
    ws = wb["dt_hscodes"]
    out: dict[str, str] = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[1] is None:
            continue
        raw = row[1]
        code = str(int(raw)) if isinstance(raw, (int, float)) else str(raw).strip()
        unit = (row[4] or "").strip() if row[4] else ""
        out[code] = unit
    wb.close()
    return out


def normalize_hs(hs: str, hs_units: dict[str, str]) -> str:
    """Strip an extra trailing zero if it's clearly an operator typo
    (e.g. '848410000' -> '84841000' if dt_hscodes has the shorter form)."""
    if hs in hs_units:
        return hs
    cand = hs
    while cand.endswith("0") and len(cand) > 6:
        cand = cand[:-1]
        if cand in hs_units:
            return cand
    return hs


# --------------------------------------------------------------------------- #
# 2. Country mapping
# --------------------------------------------------------------------------- #

COUNTRY_NAME_TO_ISO2 = {
    "AFGHANISTAN": "AF", "AUSTRALIA": "AU", "AUSTRIA": "AT",
    "BELGIUM": "BE", "BRAZIL": "BR", "CANADA": "CA",
    "CHINA": "CN", "CZECH REPUBLIC": "CZ", "CZECHIA": "CZ",
    "ECUADOR": "EC", "ESTONIA": "EE", "FRANCE": "FR",
    "GERMANY": "DE", "HUNGARY": "HU", "ICELAND": "IS",
    "INDIA": "IN", "INDONESIA": "ID", "IRELAND": "IE",
    "ITALY": "IT", "JAMAICA": "JM", "JAPAN": "JP",
    "JERSEY": "JE", "KOREA": "KR", "SOUTH KOREA": "KR",
    "KOREA, REPUBLIC OF": "KR", "MALAYSIA": "MY",
    "MEXICO": "MX", "OMAN": "OM", "PHILIPPINES": "PH",
    "POLAND": "PL", "PORTUGAL": "PT", "ROMANIA": "RO",
    "SERBIA": "RS", "SERBIA (REPUBLIC OF SERBIA)": "RS",
    "SINGAPORE": "SG", "SLOVAKIA": "SK", "SLOVENIA": "SI",
    "SOUTH AFRICA": "ZA", "SPAIN": "ES", "SWEDEN": "SE",
    "TAIWAN": "TW", "THAILAND": "TH", "TURKEY": "TR",
    "TÜRKİYE": "TR", "UNITED ARAB EMIRATES": "AE",
    "UAE": "AE", "UNITED KINGDOM": "GB", "GREAT BRITAIN": "GB",
    "UNITED STATES": "US", "USA": "US",
    "UNITED STATES MINOR OUTLYING ISLANDS": "UM",
    "VANUATU": "VU", "VIET NAM": "VN", "VIETNAM": "VN",
    "ZAMBIA": "ZM",
}


def country_to_iso2(name: str | None) -> str:
    if not name:
        return ""
    return COUNTRY_NAME_TO_ISO2.get(str(name).strip().upper(), "")


# --------------------------------------------------------------------------- #
# 3. Aggregation — from raw INV sheet OR from pre-aggregated HS CODE SUM
# --------------------------------------------------------------------------- #

def aggregate_inv(inv_ws) -> list[dict]:
    """Read a raw INVOICE sheet and aggregate by (HS, Country).

    Auto-detects columns by header label scan. Expected labels:
      'QTY' or 'SHIP QTY' or 'QUANTITY'
      'TOTAL PRICE' or 'AMOUNT USD' or 'AMOUNT'
      'WEIGHT' or 'GROSS WEIGHT'
      'COO' or 'COUNTRY OF ORIGIN' or 'COUNTRY'
      'HS CODE' or 'HSCODE' or 'HS'
    """
    label_map = {
        "qty":     ["qty", "ship qty", "quantity"],
        "amount":  ["total price", "amount usd", "amount", "amount_usd"],
        "weight":  ["weight", "gross weight"],
        "country": ["coo", "country of origin", "country"],
        "hs":      ["hs code", "hscode", "hs"],
    }
    cols: dict[str, int] = {}

    for r_idx in range(1, 30):
        for c_idx in range(1, inv_ws.max_column + 1):
            v = inv_ws.cell(row=r_idx, column=c_idx).value
            if not isinstance(v, str):
                continue
            vn = v.strip().lower()
            for key, opts in label_map.items():
                if key not in cols and vn in opts:
                    cols[key] = c_idx
        if len(cols) >= 4:
            break

    needed = {"qty", "amount", "country", "hs"}  # weight is optional
    if needed - cols.keys():
        raise RuntimeError(
            f"INV sheet: missing columns {needed - cols.keys()}. Found: {cols}"
        )

    qty_c = cols["qty"]; amt_c = cols["amount"]
    wt_c = cols.get("weight")
    ctry_c = cols["country"]; hs_c = cols["hs"]

    bucket: dict[tuple[str, str], dict] = OrderedDict()

    for row in inv_ws.iter_rows(min_row=1, values_only=True):
        hs_raw = row[hs_c - 1] if len(row) >= hs_c else None
        if hs_raw is None:
            continue
        if isinstance(hs_raw, str):
            s = hs_raw.strip()
            if not s.isdigit():
                continue
            hs_str = s
        elif isinstance(hs_raw, (int, float)):
            hs_str = str(int(hs_raw))
        else:
            continue

        country_full = row[ctry_c - 1] if len(row) >= ctry_c else None
        if not country_full:
            continue
        country_full = str(country_full).strip().upper()
        iso2 = country_to_iso2(country_full)

        def _num(v):
            return float(v) if isinstance(v, (int, float)) else 0.0

        qty = _num(row[qty_c - 1] if len(row) >= qty_c else 0)
        amt = _num(row[amt_c - 1] if len(row) >= amt_c else 0)
        wt = _num(row[wt_c - 1] if wt_c and len(row) >= wt_c else 0)

        key = (hs_str, iso2 or country_full)
        if key not in bucket:
            bucket[key] = dict(
                hs=hs_str, country_full=country_full,
                country_iso2=iso2, qty=0.0, amount=0.0, weight=0.0,
            )
        bucket[key]["qty"] += qty
        bucket[key]["amount"] += amt
        bucket[key]["weight"] += wt

    return sorted(
        bucket.values(),
        key=lambda r: (r["hs"], r["country_iso2"] or "ZZ"),
    )


def aggregate_hs_code_sum(ws, iso2_to_name: dict[str, str] | None = None) -> list[dict]:
    """Read a pre-aggregated 'HS CODE SUM' / 'HSSC' sheet.

    Expected columns (1-based, header on first non-empty row):
      A=SN, B=HS CODE, C=QTY, D=WEIGHT, E=AMOUNT, F=COO (ISO2)
    or the WENS HSSC layout:
      B=SN, C=HS CODE, D=COO_full, E=QTY, F=AMOUNT, G=WEIGHT, H=DESC, K=ISO2
    Auto-detects by inspecting the header row.
    """
    iso2_to_name = iso2_to_name or {iso2: full for full, iso2 in COUNTRY_NAME_TO_ISO2.items()}

    # Find header row & layout
    header_row = None
    layout = None
    for r_idx in range(1, min(ws.max_row + 1, 40)):
        vals = [ws.cell(row=r_idx, column=c).value for c in range(1, 13)]
        labels = [str(v).strip().upper() if v else '' for v in vals]
        if 'HS CODE' in labels and 'COO' in labels:
            header_row = r_idx
            if labels[1] == 'HS CODE':       # TPL HS CODE SUM layout
                layout = 'tpl'
            elif labels[2] == 'HS CODE':     # WENS HSSC layout
                layout = 'wens'
            break

    if header_row is None:
        raise RuntimeError("Could not locate HS CODE SUM / HSSC header")

    rows: list[dict] = []
    for r_idx in range(header_row + 1, ws.max_row + 1):
        if layout == 'tpl':
            hs = ws.cell(row=r_idx, column=2).value
            qty = ws.cell(row=r_idx, column=3).value
            wt = ws.cell(row=r_idx, column=4).value
            amt = ws.cell(row=r_idx, column=5).value
            iso2 = (ws.cell(row=r_idx, column=6).value or '').strip().upper() if isinstance(ws.cell(row=r_idx, column=6).value, str) else ''
        else:  # wens
            hs = ws.cell(row=r_idx, column=3).value
            country_full = ws.cell(row=r_idx, column=4).value
            qty = ws.cell(row=r_idx, column=5).value
            amt = ws.cell(row=r_idx, column=6).value
            wt = ws.cell(row=r_idx, column=7).value
            iso2_raw = ws.cell(row=r_idx, column=11).value
            iso2 = iso2_raw.strip().upper() if isinstance(iso2_raw, str) else ''

        if not isinstance(hs, (int, float)) or not iso2:
            continue
        # Skip placeholder rows in blank templates (HS=0, ISO2='#N/A'/'0'/'')
        if int(hs) == 0:
            continue
        if iso2 in ("#N/A", "0", "NA", "N/A", ""):
            continue

        rows.append(dict(
            hs=str(int(hs)),
            country_full=iso2_to_name.get(iso2, iso2),
            country_iso2=iso2,
            qty=float(qty or 0),
            amount=float(amt or 0),
            weight=float(wt or 0),
        ))

    rows.sort(key=lambda r: (r["hs"], r["country_iso2"]))
    return rows


# --------------------------------------------------------------------------- #
# 4. Write HSSC sheet in WENS template (locate header row dynamically)
# --------------------------------------------------------------------------- #

HSSC_DESCRIPTION = "AUTO SPARE PARTS & COMPONENTS"   # only used in HSSC sheet


def _find_hssc_data_start(ws) -> int:
    for r in range(1, min(ws.max_row + 1, 80)):
        b = ws.cell(row=r, column=2).value
        c = ws.cell(row=r, column=3).value
        if (isinstance(b, str) and b.strip().upper() == "SN" and
                isinstance(c, str) and c.strip().upper() in ("HS CODE", "HSCODE")):
            return r + 1
    return 23  # OMEGAPART-style fallback


def write_hssc_sheet(wb, agg_rows: list[dict]) -> None:
    ws = wb["HSSC"]
    start = _find_hssc_data_start(ws)
    for r in range(start, ws.max_row + 1):
        for c in range(2, 12):
            ws.cell(row=r, column=c).value = None
    for i, rec in enumerate(agg_rows, start=1):
        r = start + (i - 1)
        ws.cell(row=r, column=2, value=i)
        try:
            ws.cell(row=r, column=3, value=int(rec["hs"]))
        except ValueError:
            ws.cell(row=r, column=3, value=rec["hs"])
        ws.cell(row=r, column=4, value=rec["country_full"])
        ws.cell(row=r, column=5, value=int(round(rec["qty"])))
        ws.cell(row=r, column=6, value=round(rec["amount"], 2))
        ws.cell(row=r, column=7, value=round(rec["weight"], 3))
        ws.cell(row=r, column=8, value=HSSC_DESCRIPTION)
        ws.cell(row=r, column=11, value=rec["country_iso2"])


def write_piv_hsc_sheet(wb, agg_rows: list[dict]) -> None:
    ws = wb["PIV HSC"]
    if ws.max_row >= 1:
        for r in range(1, ws.max_row + 1):
            for c in range(1, ws.max_column + 1):
                ws.cell(row=r, column=c).value = None
    hs_list: list[str] = []
    country_list: list[str] = []
    seen_hs, seen_c = set(), set()
    for r in agg_rows:
        if r["hs"] not in seen_hs:
            seen_hs.add(r["hs"]); hs_list.append(r["hs"])
        c2 = r["country_iso2"] or r["country_full"]
        if c2 not in seen_c:
            seen_c.add(c2); country_list.append(c2)
    country_list.sort()
    ws.cell(row=1, column=2, value="HS Code")
    for idx, c2 in enumerate(country_list):
        ws.cell(row=1, column=3 + idx, value=c2)
    lookup = {(r["hs"], r["country_iso2"] or r["country_full"]): r for r in agg_rows}
    for r_idx, hs in enumerate(hs_list):
        row_no = 2 + r_idx
        try:
            ws.cell(row=row_no, column=2, value=int(hs))
        except ValueError:
            ws.cell(row=row_no, column=2, value=hs)
        for c_idx, c2 in enumerate(country_list):
            rec = lookup.get((hs, c2))
            if rec is not None:
                ws.cell(row=row_no, column=3 + c_idx,
                        value=round(rec["weight"], 3))


# --------------------------------------------------------------------------- #
# 5. TXT generation — final format calibrated against TPL2605012.txt
# --------------------------------------------------------------------------- #

ID_DESCRIPTION_TEMPLATE = "{qty} pcs AUTOSPARE PARTS"


def _fmt_g(v: float) -> str:
    """Trim trailing zeros (and the dot when whole), no scientific notation.

    Examples:
      4737.0   -> '4737'
      71.9     -> '71.9'
      135.223  -> '135.223'
      0.032    -> '0.032'
      2629.0   -> '2629'
      2194.20  -> '2194.2'   (float can't distinguish .20 from .2)
    """
    return f"{v:.10g}"


def _fmt_qty_int(q: float) -> str:
    return str(int(round(q)))


def _convert_date_to_yy(date_str: str) -> str:
    """Convert YYYY-MM-DD -> YY-MM-DD. Pass through if already short."""
    s = (date_str or '').strip()
    if len(s) == 10 and s[4] == '-' and s[7] == '-':
        return s[2:]
    return s


def _ih_row(values: list[str]) -> str:
    return ",".join('"' + v + '"' for v in values) + "\n"


def _id_row(values: list[str]) -> str:
    return ",".join('"' + v + '"' for v in values) + "\n"


def generate_txt(
    agg_rows: list[dict],
    *,
    decl_no: str,
    decl_date: str,          # accepts both YYYY-MM-DD and YY-MM-DD
    pages: int,
    company: str,            # the SUPPLIER company name
    supplier_code: int = 1,  # IH field 7 — supplier code (1, 7, ...)
    total_usd: float | None = None,
    incoterm: str = "EXW",
    hs_units: dict[str, str] | None = None,
) -> str:
    """Return the full TXT content for the customs declaration.

    Output format calibrated against TPL2605012.txt:
      IH = 15 fields
      ID = 18 fields, no '1023' field
      Date = YY-MM-DD
      Description per ID = "{qty} pcs AUTOSPARE PARTS"
      Field 6 (unit-qty): for kg/Square Meters/missing => weight,
                          for u => qty, for inv/na => "?"
      Weight & amount formatted with .10g (trimmed trailing zeros)
    """
    hs_units = hs_units or {}

    if total_usd is None:
        total_usd = sum(r["amount"] for r in agg_rows)

    out = io.StringIO()

    # IH header — exactly 15 fields per TPL reference
    ih = [
        "IH",                              # [0]
        decl_no,                           # [1]
        _convert_date_to_yy(decl_date),    # [2]  YY-MM-DD
        str(pages),                        # [3]
        "1",                               # [4]
        company,                           # [5]  SUPPLIER name
        str(supplier_code),                # [6]  supplier code (1, 7, ...)
        "1",                               # [7]
        "USD",                             # [8]
        _fmt_g(total_usd) if total_usd != int(total_usd) else f"{total_usd:.2f}",
        incoterm,                          # [10]
        "", "", "", "",                    # [11..14] empty
    ]
    # Note: total_usd is normally already 2-decimal money, so keep as e.g. "165806.64"
    ih[9] = f"{total_usd:.2f}".rstrip("0").rstrip(".") if total_usd != round(total_usd, 2) else f"{total_usd:.2f}"
    # Better: just use 2-decimal form, but trim if whole number.
    # In references all totals are written with two decimals (e.g. 165806.64, 315666.27, 279495.76).
    ih[9] = f"{total_usd:.2f}"
    out.write(_ih_row(ih))

    # ID rows — 18 fields each
    for idx, rec in enumerate(agg_rows, start=1):
        hs = normalize_hs(rec["hs"], hs_units)
        unit = (hs_units.get(hs, "") or "").strip()
        u_low = unit.lower()

        if unit == "kg":
            label, field6 = "kg", _fmt_g(rec["weight"])
        elif unit == "u":
            label, field6 = "u", _fmt_qty_int(rec["qty"])
        elif u_low == "square meters":
            label, field6 = "Square Meters", _fmt_qty_int(rec["qty"])
        elif u_low in ("inv", "na"):
            label, field6 = "?", "?"
        else:
            # Missing from dt_hscodes — default to kg (most common, matches
            # reference behavior for codes like 27101911 / 27101992 / 38190000
            # that are absent from the database but treated as kg in TXT).
            label, field6 = "kg", _fmt_g(rec["weight"])

        description = ID_DESCRIPTION_TEMPLATE.format(qty=_fmt_qty_int(rec["qty"]))

        id_row = [
            "ID",                          # [0]
            str(idx),                      # [1]
            hs,                            # [2]
            description,                   # [3]  "{qty} pcs AUTOSPARE PARTS"
            "N",                           # [4]
            label,                         # [5]  kg / u / Square Meters / ?
            field6,                        # [6]  weight or qty or ?
            "kg",                          # [7]
            _fmt_g(rec["weight"]),         # [8]  weight (trimmed)
            "", "",                        # [9..10] empty
            _fmt_g(rec["amount"]),         # [11]  amount USD (trimmed)
            rec["country_iso2"] or "",     # [12]
            "", "", "", "", "",            # [13..17] empty (NO "1023")
        ]
        out.write(_id_row(id_row))

    return out.getvalue()


# --------------------------------------------------------------------------- #
# 6. Convenience: pages calculator (formula fallback)
# --------------------------------------------------------------------------- #

ROWS_PER_PAGE = 142  # matches OMEGAPART_94 (32) and WENS_100parts (49)


def count_inv_pages(wb_or_path) -> int:
    """Estimate the number of printable pages of the INV sheet by row count.

    Use as a fallback when the user doesn't specify pages explicitly.
    Note: in the TPL2605012 reference pages=1 with 283 INV rows, which the
    formula gives as 2 — so this is a HEURISTIC, not authoritative. Prefer
    user-supplied value for accurate output.
    """
    if isinstance(wb_or_path, (str, Path)):
        wb = openpyxl.load_workbook(wb_or_path, data_only=True, read_only=True)
        own_wb = True
    else:
        wb = wb_or_path
        own_wb = False

    ws = wb["INV"]
    last_num = 0
    for r_idx, row in enumerate(ws.iter_rows(values_only=True), 1):
        if r_idx > 50_000:
            break
        if len(row) < 2:
            continue
        v = row[1]
        if isinstance(v, (int, float)) and 1 <= v < 1_000_000:
            if v > last_num:
                last_num = v
    if own_wb:
        wb.close()

    if last_num == 0:
        return 1
    return max(1, math.ceil(last_num / ROWS_PER_PAGE))
