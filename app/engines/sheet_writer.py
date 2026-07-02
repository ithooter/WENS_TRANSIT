"""
sheet_writer.py — единый «писатель таблицы» для шаблонов (ядро анти-дублирования).

Раньше логика заполнения INV и PL была продублирована в fill_inv_pl.py и
fill_optiauto.py под разные (захардкоженные) геометрии шаблонов. Здесь механика
вынесена ОДИН раз и параметризуется через `SheetGeometry`:

  • где начинаются/заканчиваются строки данных;
  • какое логическое поле пишется в какую колонку;
  • какие колонки — «формулы-справочники» строки (VLOOKUP и т.п.), которые надо
    клонировать на добавленные строки;
  • как переписать блок ИТОГО при вставке строк (callback per-template).

Только ЗНАЧЕНИЯ ячеек данных пишутся — стили/форматы шаблона не трогаются, поэтому
заполненные строки автоматически наследуют оформление. Формулы при вставке строк
сдвигаются через openpyxl Translator (корректно для относительных/абсолютных
ссылок), а не хрупким str.replace.

Работает на ЛЮБОЕ число строк: если данных больше вместимости шаблона — строки
вставляются со стилем эталонной строки; если меньше — лишние очищаются.
"""
from __future__ import annotations

from copy import copy
from dataclasses import dataclass, field
from typing import Callable, Optional

from openpyxl.formula.translate import Translator
from openpyxl.utils import get_column_letter


@dataclass
class SheetGeometry:
    """Геометрия одного листа-шаблона (INV или PL)."""
    sheet_name: str
    data_start: int                       # первая строка данных (эталон стиля)
    data_end: int                         # последняя строка данных до расширения (вкл.)
    template_cols: dict[str, int]         # логическое поле → колонка (1-based)
    style_cols: tuple[int, int]           # (min_col, max_col) — какие колонки стилизовать
    helper_formula_cols: tuple[int, ...] = ()   # колонки с формулой в data_start (клонируем через Translator)
    renumber_field: Optional[str] = None  # поле, которое переписываем как 1,2,3…
    # Переписать ИТОГО/ссылки после ВСТАВКИ строк: (ws, data_start, new_end, extra).
    on_expand: Optional[Callable[..., None]] = None

    @property
    def capacity(self) -> int:
        return self.data_end - self.data_start + 1


def _copy_style(ws, src_row: int, dst_row: int, min_col: int, max_col: int) -> None:
    for c in range(min_col, max_col + 1):
        src = ws.cell(row=src_row, column=c)
        dst = ws.cell(row=dst_row, column=c)
        dst.font = copy(src.font)
        dst.border = copy(src.border)
        dst.fill = copy(src.fill)
        dst.alignment = copy(src.alignment)
        dst.number_format = src.number_format


def _clone_helper_formulas(ws, src_row: int, dst_row: int, cols) -> None:
    """Клонировать формулы-справочники эталонной строки на новую строку, сдвигая
    относительные ссылки (S23→S24 и т.п.), сохраняя абсолютные ($B$2:$G$254)."""
    for c in cols:
        f = ws.cell(row=src_row, column=c).value
        if isinstance(f, str) and f.startswith("="):
            letter = get_column_letter(c)
            translated = Translator(f, origin=f"{letter}{src_row}").translate_formula(
                f"{letter}{dst_row}")
            ws.cell(row=dst_row, column=c).value = translated


def fill_sheet(ws, rows: list[dict], geom: SheetGeometry, *,
               constants: dict | None = None) -> dict:
    """Записать `rows` в лист по геометрии `geom`. Возвращает отчёт.

    rows: список словарей {логическое_поле: значение} (canonical → уже отображён).
    constants: {поле: значение} в каждую строку данных (напр. {'vat_rate': 0}).
    """
    constants = constants or {}
    n = len(rows)
    start = geom.data_start
    end = geom.data_end
    expanded = False

    # --- расширение под число строк больше вместимости шаблона ---
    if n > geom.capacity:
        extra = n - geom.capacity
        ws.insert_rows(end + 1, extra)          # блок ИТОГО (end+1…) уезжает вниз
        new_end = end + extra
        smin, smax = geom.style_cols
        for r in range(end + 1, new_end + 1):
            _copy_style(ws, start, r, smin, smax)
            _clone_helper_formulas(ws, start, r, geom.helper_formula_cols)
        if geom.on_expand:
            geom.on_expand(ws, start, new_end, extra)
        end = new_end
        expanded = True

    # --- запись значений (стили не трогаем) ---
    for i, rec in enumerate(rows):
        r = start + i
        for field_name, c in geom.template_cols.items():
            if field_name == geom.renumber_field:
                ws.cell(row=r, column=c).value = i + 1
            elif field_name in constants:
                ws.cell(row=r, column=c).value = constants[field_name]
            elif rec.get(field_name) not in (None, ""):
                ws.cell(row=r, column=c).value = rec[field_name]
            else:
                ws.cell(row=r, column=c).value = None

    # --- очистка неиспользованных строк данных (стили остаются) ---
    for r in range(start + n, end + 1):
        for c in geom.template_cols.values():
            ws.cell(row=r, column=c).value = None
        for c in geom.helper_formula_cols:
            ws.cell(row=r, column=c).value = None

    return {"rows_written": n, "expanded": expanded, "new_end": end}
