"""
fill_optiauto.py — fill INV + PL sheets of the OptiAuto template, keeping the
original design (header, addresses, red column headers, totals formulas).

⚠️  ВЕРСИЯ ИЗ КЭША: этот файл — версия из исходного проекта. Файл `fill_nova.py`
    вызывает fill_optiauto() с параметрами `inv_enrichment` и `pl_dedup_field`,
    которых в ЭТОЙ версии НЕТ. Для режима «заполнить шаблон» (Спринт 2) замени
    этот файл актуальной версией с поддержкой этих параметров.

Workflow (OptiAuto mode):
    source INV book  ─┐
    source PL book   ─┼─▶  write into OPTIAUTO template copy  ─▶  filled .xlsx
    OptiAuto template ┘

Design preservation rules:
  * Only data-row cell *values* are written. Cell styles (borders, number
    formats, fonts, fills) of the template are left untouched, so filled data
    inherits the template's formatting automatically.
  * Header/address rows and the red column-header rows are never modified.
  * Totals formulas (SUBTOTAL/SUM over fixed ranges) are left in place and
    recalculate when the file is opened.
  * Unused placeholder data rows are blanked (value=None) but keep their style.

Template geometry (OPTIAUTO_APMG_00_2025):
  INV: column headers on row 22 (fill FFC00000), data rows 23..577,
       totals on rows 578-580.
  PL : column headers on rows 19-20 (fill FFC00000), data rows 21..158,
       totals on rows 116-118.

Source column mapping is auto-detected by header label, so the source books
can have columns in any order as long as the headers are recognizable.
"""

from __future__ import annotations
from pathlib import Path
import shutil

import openpyxl


# --------------------------------------------------------------------------- #
# Template geometry
# --------------------------------------------------------------------------- #

INV_DATA_START = 23
INV_DATA_END = 577          # inclusive; totals formulas sum G23:G577 etc.
PL_DATA_START = 21
PL_DATA_END = 114           # data rows 21-114; row 115 blank; totals 116-118
#                             (SUM ranges extend to F158 but real data ends 114)

# Template column index (1-based) for each logical field.
INV_TEMPLATE_COLS = {
    "n":            2,   # N (sequential — we renumber)
    "brand":        3,   # BRAND
    "part_number":  4,   # PART NUMBER
    "description":  5,   # DESCRIPTION RU
    "description2": 6,   # DESCRIPTION RU (2nd column)
    "qty":          7,   # QTY
    "unit_weight":  8,   # UNIT WEIGHT
    "unit_s_price": 9,   # UNIT S.PRICE
    "unit_b_price": 10,  # UNIT B.PRICE
    "amt_b":        11,  # AMT B
    "amt_aed":      12,  # AMT AED
    "vat_rate":     14,  # VAT RATE %
    "vat_amt":      15,  # VAT AMT
    "coo":          16,  # COO
    "hs_code":      17,  # HS CODE
    "incoming_decl":18,  # INCOMING DECLARATION
}

PL_TEMPLATE_COLS = {
    "pkg_number":    2,  # Номер Коробки / PKG Number
    "width":         3,  # W
    "height":        4,  # H
    "length":        5,  # L
    "qty":           6,  # кол-во мест / QTY
    "gross_weight":  7,  # Вес / GR.W
    "cbm":           8,  # CBM(m3)
    "place":        11,  # Place
    "actual_weight":12,  # ACTUAL WEIGHT
}

# Header-label aliases for auto-detecting columns in the SOURCE books.
# Calibrated against the OptiAuto source format (sheet "Invoice List"):
#   N | Brand | Part number | Description | Description Rus | Weight |
#   Quantity | Price AED | Total AED | Origin | HSCode
INV_SOURCE_ALIASES = {
    "n":            ["n", "№", "no", "sr.no", "sr no", "sr.no."],
    "brand":        ["brand", "бренд"],
    "part_number":  ["part number", "part no", "part_number", "артикул", "номер детали"],
    # Russian description ONLY (never the English "Description" column):
    "description":  ["description russian", "description rus", "description ru",
                     "russian name", "наименование рус", "русское наименование",
                     "наименование"],
    # English description — read only as a fallback for blank Russian ones:
    "description_en": ["description", "english name", "англ наименование",
                       "description en", "name"],
    "qty":          ["quantity", "qty", "кол-во", "ship qty"],
    "unit_weight":  ["weight", "unit weight", "вес ед", "unit_weight", "вес"],
    "unit_s_price": ["price aed", "price usd", "unit s.price", "unit s price",
                     "s.price", "supplier price", "price"],
    "amt_aed":      ["total aed", "total usd", "amt aed", "amount aed",
                     "amt_aed", "amount", "total price", "total"],
    "coo":          ["origin", "coo", "country of origin", "country", "страна"],
    "hs_code":      ["hscode", "hs code", "hs", "тн вэд"],
    "incoming_decl":["incoming declaration", "incoming decl", "декларация"],
}

PL_SOURCE_ALIASES = {
    "pkg_number":   ["box n", "box no", "pkg number", "номер коробки",
                     "box number", "pkg no", "package"],
    "width":        ["width", "w", "ширина"],
    "height":       ["height", "h", "высота"],
    "length":       ["length", "l", "длина"],
    "qty":          ["qty", "кол-во мест", "кол-во", "quantity", "места"],
    "gross_weight": ["weight", "gr.w", "gross weight", "вес", "gross", "gr w"],
    "cbm":          ["cbm", "cbm(m3)", "cbm m3", "объем", "volume"],
    "place":        ["place", "место"],
    "actual_weight":["actual weight", "факт вес"],
}


# --------------------------------------------------------------------------- #
# Source reading
# --------------------------------------------------------------------------- #

def _detect_source_columns(ws, aliases: dict[str, list[str]],
                           scan_rows: int | None = 30) -> tuple[dict[str, int], int]:
    """Scan a source sheet for the header row, mapping logical field names to
    1-based column indexes. Returns (col_map, header_row_index).

    The header row is the one where the most aliases match — so when a sheet
    contains several tables (e.g. a PL with an item list AND a separate
    box-dimension table), the richer table header wins. scan_rows=None scans
    the whole sheet.
    """
    limit = ws.max_row if scan_rows is None else min(scan_rows, ws.max_row)
    best_row, best_map, best_score = None, {}, 0
    for r in range(1, limit + 1):
        col_map: dict[str, int] = {}
        for c in range(1, ws.max_column + 1):
            v = ws.cell(row=r, column=c).value
            if not isinstance(v, str):
                continue
            label = v.strip().lower()
            for field, opts in aliases.items():
                if field not in col_map and label in opts:
                    col_map[field] = c
        if len(col_map) > best_score:
            best_score, best_row, best_map = len(col_map), r, col_map
    return best_map, (best_row or 1)


def _read_source_rows(ws, col_map: dict[str, int], header_row: int) -> list[dict]:
    """Read data rows below the header into a list of {field: value} dicts.

    A row is data if it has any non-empty value in the mapped columns and a
    usable key field (part_number / pkg_number / qty / hs_code).

    Summary/total rows are skipped, but the summary-word check is applied ONLY
    to label columns (n / pkg_number / brand) — never to data columns such as
    COO, so a country like "NETHERLANDS" is not mistaken for a "net..." total.
    """
    rows: list[dict] = []
    key_fields = [f for f in ("part_number", "pkg_number", "qty", "hs_code")
                  if f in col_map]
    SUMMARY_PREFIXES = ("total", "итого", "subtotal", "всего", "grand total",
                        "net ", "нетто", "container", "контейнер", "gross ")
    LABEL_FIELDS = ("n", "pkg_number", "brand")   # only these can flag a summary
    for r in range(header_row + 1, ws.max_row + 1):
        rec = {}
        any_val = False
        is_summary = False
        for field, c in col_map.items():
            v = ws.cell(row=r, column=c).value
            rec[field] = v
            if v not in (None, ""):
                any_val = True
                if (field in LABEL_FIELDS and isinstance(v, str)
                        and v.strip().lower().startswith(SUMMARY_PREFIXES)):
                    is_summary = True
        if not any_val or is_summary:
            continue
        if key_fields and all(rec.get(f) in (None, "") for f in key_fields):
            continue
        rows.append(rec)
    return rows


# --------------------------------------------------------------------------- #
# Filling (value-only, preserves styles)
# --------------------------------------------------------------------------- #

def _expand_inv_capacity(ws, needed_rows: int) -> int:
    """Insert rows into the INV data area when the source exceeds the template
    capacity, copying styles + the per-row COUNTRY VLOOKUP, and rewriting the
    totals formulas + header references. Returns the new data-end row.

    Does nothing (returns INV_DATA_END) if the data already fits.
    """
    from copy import copy
    capacity = INV_DATA_END - INV_DATA_START + 1
    if needed_rows <= capacity:
        return INV_DATA_END

    extra = needed_rows - capacity
    ws.insert_rows(INV_DATA_END + 1, extra)      # push totals (578+) down
    new_end = INV_DATA_END + extra

    # Style only the NEWLY inserted rows (23..577 already carry template style).
    style_src = INV_DATA_START
    src_cells = [ws.cell(row=style_src, column=c) for c in range(1, 20)]
    for r in range(INV_DATA_END + 1, new_end + 1):
        for ci, c in enumerate(range(1, 20)):
            src = src_cells[ci]
            dst = ws.cell(row=r, column=c)
            dst.font = copy(src.font)
            dst.border = copy(src.border)
            dst.fill = copy(src.fill)
            dst.number_format = src.number_format
            dst.alignment = copy(src.alignment)
    # VLOOKUP (country alpha-2) on rows that may lack it (575..new_end).
    for r in range(575, new_end + 1):
        ws.cell(row=r, column=19).value = (
            f"=VLOOKUP(P{r},'COUNTRY LIST1'!$C$2:$D$248,2,0)")

    # Rewrite totals (originally rows 578/579/580, now shifted by `extra`).
    t1, t2, t3 = 578 + extra, 579 + extra, 580 + extra
    s, e = INV_DATA_START, new_end
    ws.cell(row=t1, column=7).value  = f"=SUBTOTAL(9,G{s}:G{e})"
    ws.cell(row=t1, column=11).value = f"=SUM(K{s}:K{e})"
    ws.cell(row=t1, column=12).value = f"=SUM(L{s}:L{e})"
    ws.cell(row=t1, column=13).value = f"=SUM(M{s}:M{e})"
    ws.cell(row=t1, column=14).value = f"=SUM(N{s}:N{e})"
    ws.cell(row=t2, column=15).value = f"=SUM(O{s}:O{e})"
    ws.cell(row=t3, column=11).value = f"=K{t1}"
    ws.cell(row=t3, column=12).value = f"=L{t1}"
    ws.cell(row=t3, column=13).value = f"=M{t1}+O{t2}"

    # Fix header references that pointed at the old totals rows.
    ws.cell(row=16, column=14).value = f"=G{t1}"   # Total Number of Units
    ws.cell(row=17, column=14).value = f"=L{t3}"   # Total Amount (incl VAT)

    return new_end


def _expand_pl_capacity(ws, needed_rows: int, inv_ws=None) -> int:
    """Insert rows into the PL data area when the source exceeds capacity,
    copying styles + the per-row =L-G helper, and rewriting the totals SUM
    ranges. Also fixes INV header references (=PL!C116/117/118) if inv_ws is
    given. Returns the new data-end row. No-op if it already fits.
    """
    from copy import copy
    capacity = PL_DATA_END - PL_DATA_START + 1
    if needed_rows <= capacity:
        return PL_DATA_END

    extra = needed_rows - capacity
    ws.insert_rows(PL_DATA_END + 1, extra)        # insert after last data row
    new_end = PL_DATA_END + extra

    style_src = PL_DATA_START
    src_cells = [ws.cell(row=style_src, column=c) for c in range(1, 15)]
    for r in range(PL_DATA_END + 1, new_end + 1):
        for ci, c in enumerate(range(1, 15)):
            src = src_cells[ci]
            dst = ws.cell(row=r, column=c)
            dst.font = copy(src.font)
            dst.border = copy(src.border)
            dst.fill = copy(src.fill)
            dst.number_format = src.number_format
            dst.alignment = copy(src.alignment)
        ws.cell(row=r, column=14).value = f"=L{r}-G{r}"   # ACTUAL - GROSS helper

    # Totals were at 116/117/118; they shifted down by `extra`.
    s, e = PL_DATA_START, new_end
    tot_pkg, tot_wt, tot_vol = 116 + extra, 117 + extra, 118 + extra
    ws.cell(row=tot_pkg, column=3).value = f"=SUM(F{s}:F{e})"   # Total PKG's
    ws.cell(row=tot_wt,  column=3).value = f"=SUM(G{s}:G{e})"   # Total Weight
    ws.cell(row=tot_vol, column=3).value = f"=SUM(H{s}:H{e})"   # Total Volume

    # Fix INV header references that point at the PL totals.
    if inv_ws is not None:
        inv_ws.cell(row=18, column=17).value = f"=PL!C{tot_pkg}"  # Num of Pckg's
        inv_ws.cell(row=19, column=17).value = f"=PL!C{tot_vol}"  # Total Vol CBM
        inv_ws.cell(row=20, column=17).value = f"=PL!C{tot_wt}"   # Total Weight
    return new_end


def _write_rows(ws, rows: list[dict], template_cols: dict[str, int],
                start_row: int, end_row: int, *,
                renumber_field: str | None = None,
                constants: dict | None = None) -> int:
    """Write source rows into template data area. Returns rows written.

    Only cell .value is set, so template cell styles are preserved. Unused
    rows (start..end) are blanked. Raises if rows exceed capacity.

    constants: {logical_field: value} written to every data row (after source
    data), e.g. {'vat_rate': 0} or {'qty': 1}.
    """
    constants = constants or {}
    capacity = end_row - start_row + 1
    if len(rows) > capacity:
        raise ValueError(
            f"Source has {len(rows)} rows but template capacity is {capacity} "
            f"(rows {start_row}-{end_row}). Need to insert rows + copy styles."
        )

    for i, rec in enumerate(rows):
        r = start_row + i
        for field, c in template_cols.items():
            if field == renumber_field:
                ws.cell(row=r, column=c).value = i + 1     # sequential N
            elif field in constants:
                ws.cell(row=r, column=c).value = constants[field]
            elif field in rec and rec[field] not in (None, ""):
                ws.cell(row=r, column=c).value = rec[field]
            else:
                ws.cell(row=r, column=c).value = None

    # Blank the unused placeholder rows (keep styles intact)
    for r in range(start_row + len(rows), end_row + 1):
        for c in template_cols.values():
            ws.cell(row=r, column=c).value = None

    return len(rows)


# --------------------------------------------------------------------------- #
# Public entry point
# --------------------------------------------------------------------------- #

def fill_optiauto(
    *,
    template_path: str | Path,
    source_inv_path: str | Path,
    source_pl_path: str | Path,
    output_path: str | Path,
    source_inv_sheet: str | None = None,
    source_pl_sheet: str | None = None,
    inv_constants: dict | None = None,
    pl_constants: dict | None = None,
    inv_header_updates: dict | None = None,
    blank_desc_fill: str | None = None,
    fill_pl: bool = True,
    pl_grand_totals: dict | None = None,
) -> dict:
    """Fill INV + PL of the OptiAuto template from two source books.

    source_inv_sheet / source_pl_sheet: sheet names to read when the source
    book has multiple sheets (e.g. one workbook with 'Invoice List' and
    'Packing List'). Defaults to the active sheet.

    inv_constants / pl_constants: {logical_field: constant_value} written to
    every data row (e.g. {'vat_rate': 0, 'vat_amt': 0} for INV; {'qty': 1}
    for PL where each package is one place).

    inv_header_updates: {(row, col): value} cells to set in the INV header
    (e.g. {(14, 14): date, (15, 14): invoice_no}). Styles preserved.
    """
    shutil.copy(template_path, output_path)
    wb = openpyxl.load_workbook(output_path)

    def _pick_sheet(wbk, name):
        if name and name in wbk.sheetnames:
            return wbk[name]
        for cand in ("Invoice List", "TDSheet", "INVOICE", "INV",
                     "Packing List", "Sheet1", "PL"):
            if name is None and cand in wbk.sheetnames:
                return wbk[cand]
        return wbk.active

    # ---- INV header updates (Date, Invoice No, ...) ----
    if inv_header_updates:
        inv_ws = wb["INV"]
        for (r, c), val in inv_header_updates.items():
            inv_ws.cell(row=r, column=c).value = val

    # ---- INV ----
    src_inv_wb = openpyxl.load_workbook(source_inv_path, data_only=True)
    src_inv_ws = _pick_sheet(src_inv_wb, source_inv_sheet)
    inv_cols, inv_hdr = _detect_source_columns(src_inv_ws, INV_SOURCE_ALIASES)
    inv_rows = _read_source_rows(src_inv_ws, inv_cols, inv_hdr)
    # Fill blank Russian descriptions: fall back to the English description,
    # then to a generic label, so no DESCRIPTION RU cell is left empty.
    blank_desc_filled = 0
    for rec in inv_rows:
        if not rec.get("description") or str(rec.get("description")).strip() == "":
            fallback = rec.get("description_en")
            if not fallback or str(fallback).strip() == "":
                fallback = blank_desc_fill or "Автозапчасть"
            rec["description"] = fallback
            blank_desc_filled += 1
    inv_end = _expand_inv_capacity(wb["INV"], len(inv_rows))
    inv_written = _write_rows(
        wb["INV"], inv_rows, INV_TEMPLATE_COLS,
        INV_DATA_START, inv_end, renumber_field="n",
        constants=inv_constants,
    )
    src_inv_wb.close()

    # ---- PL ----
    pl_written = 0
    pl_cols, pl_hdr = {}, None
    if fill_pl:
        src_pl_wb = openpyxl.load_workbook(source_pl_path, data_only=True)
        src_pl_ws = _pick_sheet(src_pl_wb, source_pl_sheet)
        pl_cols, pl_hdr = _detect_source_columns(src_pl_ws, PL_SOURCE_ALIASES,
                                                  scan_rows=None)
        pl_rows = _read_source_rows(src_pl_ws, pl_cols, pl_hdr)
        pl_end = _expand_pl_capacity(wb["PL"], len(pl_rows), inv_ws=wb["INV"])
        pl_written = _write_rows(
            wb["PL"], pl_rows, PL_TEMPLATE_COLS,
            PL_DATA_START, pl_end, renumber_field="pkg_number",
            constants=pl_constants,
        )
        src_pl_wb.close()
    elif pl_grand_totals is not None:
        # Source PL has no per-box dimensions: blank the box table and write
        # the shipment grand totals straight into the PL total cells so the
        # INV header references (Num Pckg's / Weight / CBM) stay correct.
        pl_ws = wb["PL"]
        for r in range(PL_DATA_START, PL_DATA_END + 1):
            for c in list(PL_TEMPLATE_COLS.values()) + [14]:
                pl_ws.cell(row=r, column=c).value = None
        if "packages" in pl_grand_totals:
            pl_ws.cell(row=116, column=3).value = pl_grand_totals["packages"]
        if "weight" in pl_grand_totals:
            pl_ws.cell(row=117, column=3).value = pl_grand_totals["weight"]
        if "cbm" in pl_grand_totals:
            pl_ws.cell(row=118, column=3).value = pl_grand_totals["cbm"]

    wb.save(output_path)
    wb.close()

    return dict(
        output=str(output_path),
        inv_rows=inv_written,
        pl_rows=pl_written,
        blank_desc_filled=blank_desc_filled,
        inv_columns_detected=inv_cols,
        inv_header_row=inv_hdr,
        pl_columns_detected=pl_cols,
        pl_header_row=pl_hdr,
    )
