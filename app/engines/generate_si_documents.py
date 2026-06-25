"""
generate_si_documents.py — автогенератор документов на отгрузку.

Поддерживает два профиля источника (автодетект по колонкам):
  - autoseller    (OPTIAUTO FZE → LLC Autoseller, валюта AED)
  - fa_logistics  (Diesel Technical → FA-Logistics, валюта USD)

Валюта берётся из источника как есть, без конвертации. НДС не применяется
(включается флагом в профиле при необходимости).

Использование:
    python3 generate_si_documents.py <source.xls|.xlsx> [output_dir]
"""

import sys, os, re, subprocess, shutil, tempfile
from pathlib import Path
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from classifier import classify_row, GROUP_ORDER

# ===================== Профили источников =====================
PROFILES = {
    'autoseller': {
        'sender': [
            'OPTIAUTO FZE',
            'Warehouse No B-07',
            'Umm Al Quwain Free Trade Zone,',
            'Al Barqaa, Umm Al Quwain, UAE',
            'optiautofze@gmail.com',
        ],
        'default_consignee': [
            'LLC «Autoseller»',
            'Room 17, Building 1A/1, Lipkinskoye Highway,',
            'Nagornoye Village, Mytishchi Urban District,',
            'Moscow Region, 141031, Russia',
        ],
        'currency': 'AED',
    },
    'fa_logistics': {
        'sender': [
            'Diesel Technical',
            '[адрес поставщика — placeholder]',
            '[город, страна]',
            '[контакт]',
        ],
        'default_consignee': [
            'FA-Logistics',
            '[адрес получателя — placeholder]',
            '[город, индекс]',
            '[страна]',
        ],
        'currency': 'USD',
    },
    'avr': {
        'sender': [
            '[Поставщик — placeholder]',
            '[адрес поставщика — placeholder]',
            '[город, страна]',
            '[контакт]',
        ],
        'default_consignee': [
            'LLC AVR',
            '[адрес получателя — placeholder]',
            '[город, индекс]',
            '[страна]',
        ],
        'currency': 'USD',
    },
    'piv_hsc': {
        'sender': [
            '[Поставщик — placeholder]',
            '[адрес поставщика — placeholder]',
            '[город, страна]',
            '[контакт]',
        ],
        'default_consignee': [
            '[Получатель — placeholder]',
            '[адрес получателя — placeholder]',
            '[город, индекс]',
            '[страна]',
        ],
        'currency': 'USD',
    },
}


def _find_soffice():
    """Ищет LibreOffice/soffice в PATH и в стандартных путях macOS/Windows/Linux."""
    candidates = [
        shutil.which('soffice'),
        shutil.which('libreoffice'),
        '/Applications/LibreOffice.app/Contents/MacOS/soffice',            # macOS
        '/usr/bin/soffice', '/usr/bin/libreoffice',                        # Linux
        '/usr/local/bin/soffice', '/opt/homebrew/bin/soffice',             # Linux/Homebrew
        r'C:\Program Files\LibreOffice\program\soffice.exe',               # Windows
        r'C:\Program Files (x86)\LibreOffice\program\soffice.exe',
    ]
    for c in candidates:
        if c and os.path.exists(c):
            return c
    return None


def to_xlsx_if_needed(src: Path) -> Path:
    """Готовит файл к чтению pandas.
    .xlsx — возвращаем как есть.
    .xls/.xlsb — конвертируем через LibreOffice, если он найден (надёжнее всего).
    Если LibreOffice не установлен — возвращаем исходный файл и полагаемся на
    pandas-движки: xlrd (для .xls) и pyxlsb (для .xlsb). Тогда эти пакеты должны
    быть установлены: python3 -m pip install xlrd pyxlsb
    """
    if src.suffix.lower() == '.xlsx':
        return src
    soffice = _find_soffice()
    if soffice:
        out_dir = Path(tempfile.gettempdir()) / 'si_convert'
        out_dir.mkdir(exist_ok=True)
        try:
            subprocess.run([soffice, '--headless', '--convert-to', 'xlsx',
                            '--outdir', str(out_dir), str(src)],
                           check=True, stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL)
            converted = out_dir / (src.stem + '.xlsx')
            if converted.exists():
                return converted
        except Exception:
            pass  # не вышло конвертировать — попробуем прочитать напрямую
    # LibreOffice не найден / не сработал — читаем исходный файл напрямую через pandas
    return src


def detect_profile(inv_columns) -> str:
    cols = {str(c).strip().lower() for c in inv_columns}
    if 'total aed' in cols: return 'autoseller'
    if 'total usd' in cols: return 'fa_logistics'  # уточнится после прочтения Details
    raise RuntimeError(f'Не удалось определить профиль. Колонки: {list(inv_columns)}')


def _read_pl_sheet(src_xlsx):
    """Считает число коробок (мест) из листа PL/Packing List.
    Поддерживает заголовки 'Box N', 'PKG Number'/'Номер Коробки', 'Colli'."""
    try:
        xl = pd.ExcelFile(src_xlsx)
    except Exception:
        return 0
    pl_sheet = next((s for s in xl.sheet_names if s.strip().upper() in ('PL', 'PACKING LIST', 'PACKINGLIST')), None)
    if pl_sheet is None:
        return 0
    raw = pd.read_excel(src_xlsx, sheet_name=pl_sheet, header=None)
    # найти строку заголовка
    hdr = None
    for i in range(min(40, len(raw))):
        vals = [str(v).strip().lower() for v in raw.iloc[i].tolist() if pd.notna(v)]
        if any(k in vals for k in ['box n', 'pkg number', 'номер коробки', 'colli']):
            hdr = i; break
    if hdr is None:
        return 0
    # первый столбец после заголовка — номера коробок; считаем непустые числовые/PACKAGE
    count = 0
    for i in range(hdr + 1, len(raw)):
        v = raw.iloc[i, 0]
        if pd.isna(v):
            # иногда номер коробки во 2-м столбце
            v = raw.iloc[i, 1] if raw.shape[1] > 1 else None
        if pd.isna(v):
            continue
        s = str(v).strip()
        if re.match(r'^\d+(\.\d+)?$', s) or re.match(r'^\s*(PACKAGE|PALLET)\s', s, re.I):
            count += 1
    return count


def _read_piv_hsc_format(src_xlsx):
    """Формат INV+PIV HSC (тип DOC_FOR_WAY), независимо от компании.
    Данные ВСЕГДА берутся из листа INV (Invoice List) и PL (Packing List);
    лист PIV HSC не используется (он не всегда заполнен).

    Служебно-задвоенные строки в INV (встречаются в некоторых выгрузках) отсеиваются
    фильтром по весу: строки без UNIT WEIGHT игнорируются.
    Реквизиты — из шапки INV (BILL TO / TAX Invoice / Date / Currency).
    Число мест — из листа PL.
    """
    profile = 'piv_hsc'
    raw_full = pd.read_excel(src_xlsx, sheet_name='INV', header=None)

    # --- найти строку заголовка таблицы INV ---
    hdr = None
    for i in range(min(40, len(raw_full))):
        vals = [str(v).strip().upper() for v in raw_full.iloc[i].tolist() if pd.notna(v)]
        if 'BRAND' in vals and any('PART NUMBER' in v for v in vals):
            hdr = i; break
    if hdr is None:
        hdr = 18

    # --- реквизиты из шапки ---
    inv_no, date_val, num_pkg, total_weight, currency = None, '', None, None, None
    consignee = []
    for i in range(hdr):
        row = raw_full.iloc[i].tolist()
        for j in range(len(row)):
            label = str(row[j]).strip() if pd.notna(row[j]) else ''
            L = label.lower()
            valnext = None
            for k in range(j + 1, len(row)):
                if pd.notna(row[k]): valnext = row[k]; break
            if 'tax invoice number' in L and valnext is not None:
                inv_no = str(valnext).strip()
            elif L == 'date' and valnext is not None:
                date_val = valnext
            elif ('num. of pckg' in L or 'num of pckg' in L) and isinstance(valnext, (int, float)):
                num_pkg = int(valnext)
            elif l_starts_weight(L) and isinstance(valnext, (int, float)):
                total_weight = float(valnext)
            elif L.startswith('currency') and valnext is not None:
                currency = str(valnext).strip()
            elif L == 'bill to:':
                for r2 in range(i + 1, min(i + 6, hdr)):
                    cv = raw_full.iloc[r2, j]
                    if pd.notna(cv) and str(cv).strip():
                        consignee.append(str(cv).strip())

    if isinstance(date_val, pd.Timestamp) or hasattr(date_val, 'strftime'):
        try: date_str = date_val.strftime('%B %-d, %Y')
        except: date_str = str(date_val)
    else:
        date_str = str(date_val).split(' ')[0] if date_val else ''

    # --- таблица INV ---
    inv = pd.read_excel(src_xlsx, sheet_name='INV', header=hdr)
    inv.columns = [str(c).strip() for c in inv.columns]
    qty_ok = inv.get('QTY', pd.Series([None] * len(inv))).apply(lambda x: isinstance(x, (int, float)) and pd.notna(x))
    inv_data = inv[qty_ok].reset_index(drop=True)
    # Фильтр служебно-задвоенных строк: оставляем только с заполненным UNIT WEIGHT
    if 'UNIT WEIGHT' in inv_data.columns:
        uw = pd.to_numeric(inv_data['UNIT WEIGHT'], errors='coerce').fillna(0)
        if (uw > 0).sum() > 0:
            inv_data = inv_data[uw > 0].reset_index(drop=True)

    amt_col = next((c for c in ['AMT S', 'AMT AED', 'AMT', 'AMT B']
                    if c in inv_data.columns and pd.to_numeric(inv_data[c], errors='coerce').fillna(0).sum() != 0), None)
    if amt_col is None:
        amt_col = next((c for c in ['AMT S', 'AMT AED', 'AMT', 'AMT B'] if c in inv_data.columns), 'AMT AED')
    desc_col = 'DESCRIPTION RU' if 'DESCRIPTION RU' in inv_data.columns else \
               'DESCRIPTION' if 'DESCRIPTION' in inv_data.columns else None

    out = pd.DataFrame()
    out['N'] = range(1, len(inv_data) + 1)
    out['Brand'] = inv_data.get('BRAND', '')
    out['Part_number'] = inv_data.get('PART NUMBER', '')
    out['Description'] = inv_data['DESCRIPTION'] if 'DESCRIPTION' in inv_data.columns else ''
    out['Description_RU'] = inv_data[desc_col] if desc_col else ''
    out['Quantity'] = pd.to_numeric(inv_data['QTY'], errors='coerce').fillna(0).astype(int)
    if 'TOTAL WEIGHT' in inv_data.columns and pd.to_numeric(inv_data['TOTAL WEIGHT'], errors='coerce').fillna(0).sum() > 0:
        out['Gross'] = pd.to_numeric(inv_data['TOTAL WEIGHT'], errors='coerce').fillna(0.0)
    else:
        uw = pd.to_numeric(inv_data.get('UNIT WEIGHT', pd.Series([0.0] * len(inv_data))), errors='coerce').fillna(0.0)
        out['Gross'] = uw.values * out['Quantity'].values
    out['Amount'] = pd.to_numeric(inv_data[amt_col], errors='coerce').fillna(0.0)
    out['Origin'] = inv_data.get('COO', '')
    out['HSCode'] = pd.to_numeric(inv_data.get('HS CODE'), errors='coerce').astype('Int64').astype(str).replace('<NA>', '')
    out['Weight_unit'] = (out['Gross'] / out['Quantity'].replace(0, 1)).round(6)
    out['Price'] = (out['Amount'] / out['Quantity'].replace(0, 1)).round(4)
    out['Currency'] = currency or 'AED'
    out['BoxN'] = ''

    # --- число мест: PL-лист → шапка Num Pckg's → 0 ---
    total_places = _read_pl_sheet(src_xlsx) or num_pkg or 0
    pl = pd.DataFrame({
        'Box':    [f'PACKAGE # {i+1}' for i in range(total_places)],
        'Width':  [None] * total_places, 'Height': [None] * total_places,
        'Length': [None] * total_places, 'Weight': [None] * total_places,
        'CBM':    [None] * total_places,
    })

    PROFILES['piv_hsc']['currency'] = currency or 'AED'
    if any('autoseller' in str(c).lower() for c in consignee):
        profile = 'autoseller'
        PROFILES['autoseller']['currency'] = currency or 'AED'

    details = {
        'invoice_no': inv_no or 'NO_NUMBER',
        'date': date_str,
        'consignee': consignee,
        'contract': '',
        'incoterms': '',
        '_override_gross': total_weight,
    }
    return out, pl, details, profile


def l_starts_weight(L):
    return L.startswith('total weight') or L.startswith('total weight k')


def _read_tdsheet_inline_format(src_xlsx):
    """Формат с одним листом TDSheet, у которого inline-шапка (Invoice No, Date),
    заголовок таблицы (строка с 'N' и 'Brand'), и сами данные. Имеет Box N в инвойсе."""
    raw = pd.read_excel(src_xlsx, sheet_name='TDSheet', header=None)

    # найти строку заголовка таблицы — там есть 'Brand' и 'Part number'
    header_row = None
    for i in range(min(30, len(raw))):
        vals = [str(v).strip() for v in raw.iloc[i].tolist() if pd.notna(v)]
        if 'Brand' in vals and 'Part number' in vals:
            header_row = i; break
    if header_row is None:
        raise RuntimeError('Не найдена строка заголовка таблицы в TDSheet')

    # Реквизиты из верхней части
    header_text = '\n'.join(str(v).strip() for v in raw.iloc[:header_row].values.flatten() if pd.notna(v))
    inv_no = _grep(r'Invoice\s*No\.?:?\s*(\S+)', header_text) or ''

    # Дата: предпочитаем ячейки, СОДЕРЖИМОЕ КОТОРЫХ — это и есть дата (без префиксов).
    # Это устойчиво к плейсхолдерам типа "Date: July 16, 2019" внутри текста шаблона.
    date = ''
    pat_month = re.compile(r'^[A-Z][a-z]+ \d{1,2},\s*\d{4}$')
    pat_dot   = re.compile(r'^\d{1,2}\.\d{1,2}\.\d{2,4}$')
    pat_slash = re.compile(r'^\d{1,2}/\d{1,2}/\d{2,4}$')
    for i in range(header_row):
        for v in raw.iloc[i].tolist():
            if not isinstance(v, str): continue
            s = v.strip()
            if pat_month.match(s) or pat_dot.match(s) or pat_slash.match(s):
                date = s; break
        if date: break
    if not date:
        m = re.search(r'([A-Z][a-z]+ \d{1,2},\s*\d{4})', header_text)
        if m: date = m.group(1)
    if not date:
        m = re.search(r'Date:?\s*(\d{1,2}[./]\d{1,2}[./]\d{2,4})', header_text)
        if m: date = m.group(1)

    # Заголовок и данные
    inv = pd.read_excel(src_xlsx, sheet_name='TDSheet', header=header_row)
    inv.columns = [str(c).strip() for c in inv.columns]
    # фильтр Total и пустых
    brand = inv.get('Brand', pd.Series([''] * len(inv)))
    has_pn = inv.get('Part number', pd.Series([None] * len(inv))).notna()
    mask = has_pn & ~brand.astype(str).str.strip().str.lower().eq('total')
    inv = inv[mask].reset_index(drop=True)
    if 'N' not in inv.columns:
        inv.insert(0, 'N', range(1, len(inv) + 1))
    inv['N'] = range(1, len(inv) + 1)

    profile = detect_profile(inv.columns)
    out = pd.DataFrame()
    out['N']           = inv['N'].astype(int)
    out['Brand']       = inv.get('Brand', '')
    out['Part_number'] = inv.get('Part number', '')
    out['Description'] = inv.get('Description', '')
    out['Description_RU'] = inv.get('Description Rus', '')
    out['Quantity']    = inv['Quantity'].astype(int)
    out['Weight_unit'] = inv['Weight']
    out['Gross']       = inv['Weight'] * inv['Quantity']
    if profile == 'autoseller':
        out['Price'] = inv['Price AED']; out['Amount'] = inv['Total AED']
    else:
        out['Price'] = inv['Price USD']; out['Amount'] = inv['Total USD']
    out['Origin']   = inv.get('Origin', '')
    out['HSCode']   = pd.to_numeric(inv['HSCode'], errors='coerce').astype('Int64').astype(str).replace('<NA>', '')
    out['Currency'] = PROFILES[profile]['currency']
    out['BoxN'] = ''
    # Box N — для точного подсчёта мест
    out['BoxN'] = inv.get('Box N', '')

    # Packing List: если есть колонка Box N в инвойсе — считаем уникальные коробки
    if 'Box N' in inv.columns:
        unique_boxes = inv['Box N'].dropna().astype(str).str.strip()
        unique_boxes = unique_boxes[unique_boxes != ''].unique()
        pl = pd.DataFrame({
            'Box':    [f'PACKAGE # {b}' for b in unique_boxes],
            'Width':  [None] * len(unique_boxes),
            'Height': [None] * len(unique_boxes),
            'Length': [None] * len(unique_boxes),
            'Weight': [None] * len(unique_boxes),
            'CBM':    [None] * len(unique_boxes),
        })
    else:
        pl = pd.DataFrame(columns=['Box', 'Width', 'Height', 'Length', 'Weight', 'CBM'])

    details = {
        'invoice_no': inv_no.strip() or 'NO_NUMBER',
        'date': date.strip(),
        'consignee': [],
        'contract': '',
        'incoterms': '',
    }
    return out, pl, details, profile


def _read_tdsheet_format(src_xlsx, xl):
    """Формат с TDSheet (1С-выгрузка): TDSheet=invoice, Sheet1=реквизиты+PL."""
    raw_inv = pd.read_excel(src_xlsx, sheet_name='TDSheet', header=0)
    raw_inv.columns = [str(c).strip() for c in raw_inv.columns]
    if 'N' not in raw_inv.columns:
        raw_inv.insert(0, 'N', range(1, len(raw_inv) + 1))
    # фильтр Total
    brand = raw_inv.get('Brand', pd.Series([''] * len(raw_inv)))
    mask = raw_inv['N'].apply(lambda x: isinstance(x, (int, float)) and pd.notna(x)) \
           & ~brand.astype(str).str.strip().str.lower().eq('total')
    raw_inv = raw_inv[mask].reset_index(drop=True)
    raw_inv['N'] = range(1, len(raw_inv) + 1)

    profile = detect_profile(raw_inv.columns)
    out = pd.DataFrame()
    out['N']           = raw_inv['N'].astype(int)
    out['Brand']       = raw_inv.get('Brand', '')
    out['Part_number'] = raw_inv.get('Part number', '')
    out['Description'] = raw_inv.get('Description', '')
    _ru = 'Description Rus' if 'Description Rus' in raw_inv.columns else \
          'Description Russian' if 'Description Russian' in raw_inv.columns else None
    out['Description_RU'] = raw_inv[_ru] if _ru else ''
    out['Quantity']    = raw_inv['Quantity'].astype(int)
    out['Weight_unit'] = raw_inv['Weight']
    out['Gross']       = raw_inv['Total Weight'] if 'Total Weight' in raw_inv.columns else raw_inv['Weight'] * raw_inv['Quantity']
    if profile == 'autoseller':
        out['Price'] = raw_inv['Price AED']; out['Amount'] = raw_inv['Total AED']
    else:
        out['Price'] = raw_inv['Price USD']; out['Amount'] = raw_inv['Total USD']
    out['Origin']   = raw_inv.get('Origin', '')
    out['HSCode']   = pd.to_numeric(raw_inv['HSCode'], errors='coerce').astype('Int64').astype(str).replace('<NA>', '')
    out['Currency'] = PROFILES[profile]['currency']
    out['BoxN'] = ''

    # Реквизиты и Packing List ищем по СОДЕРЖИМОМУ листов (кроме TDSheet),
    # потому что в разных файлах они на разных листах (Sheet1 / Sheet2 / PL).
    inv_no, date, consignee = None, None, []
    pl = pd.DataFrame(columns=['Box', 'Width', 'Height', 'Length', 'Weight', 'CBM'])
    total_packages = None

    for sheet in xl.sheet_names:
        if sheet == 'TDSheet':
            continue
        sh = pd.read_excel(src_xlsx, sheet_name=sheet, header=None)
        if sh.empty:
            continue
        # 1) Packing List — лист с заголовком 'Box N'
        first_cells = [str(v).strip() for v in sh.iloc[0].tolist() if pd.notna(v)]
        if 'Box N' in first_cells:
            sh_pl = pd.read_excel(src_xlsx, sheet_name=sheet, header=0)
            sh_pl.columns = [str(c).strip() for c in sh_pl.columns]
            mask = sh_pl['Box N'].apply(lambda x: isinstance(x, str) and re.match(r'^\s*PACKAGE\s*#', x, re.I) is not None)
            r = sh_pl[mask].reset_index(drop=True)
            pl = pd.DataFrame({
                'Box': r['Box N'], 'Width': r.get('Width'), 'Height': r.get('Height'),
                'Length': r.get('Length'), 'Weight': r.get('Weight'), 'CBM': r.get('CBM'),
            })
            continue
        # 2) Реквизиты / метки
        text = '\n'.join(str(v).strip() for v in sh.values.flatten() if pd.notna(v))
        if inv_no is None:
            inv_no = _grep(r'Tax Invoice:\s*(\S+)', text) or _grep(r'Invoice\s*No\.?:?\s*(\S+)', text)
        if date is None:
            date = _grep(r'Invoice\s*Date:?\s*(\S+)', text) or _grep(r'Date:\s*(\S+)', text)
        if not consignee:
            consignee = _block_lines(text, r'LLC|ООО|Получатель|Autoseller|AVR')
        # TOTAL NUMBER OF PACKING UNITS / TOTAL PACKAGES
        if total_packages is None:
            for _, row in sh.iterrows():
                joined = ' '.join(str(v).strip() for v in row if pd.notna(v)).upper()
                if 'TOTAL NUMBER OF PACKING UNITS' in joined or 'TOTAL PACKAGES' in joined:
                    nums = [v for v in row if isinstance(v, (int, float)) and pd.notna(v)]
                    if nums:
                        total_packages = int(nums[0]); break

    # Если packing list-таблицы не было, но есть число коробок из метки — псевдо-PL
    if pl.empty and total_packages:
        pl = pd.DataFrame({
            'Box':    [f'PACKAGE # {i+1}' for i in range(total_packages)],
            'Width':  [None] * total_packages, 'Height': [None] * total_packages,
            'Length': [None] * total_packages, 'Weight': [None] * total_packages,
            'CBM':    [None] * total_packages,
        })

    details = {
        'invoice_no': (inv_no or 'NO_NUMBER').strip() if inv_no else 'NO_NUMBER',
        'date': (date or '').strip(),
        'consignee': consignee,
        'contract': '',
        'incoterms': '',
    }
    profile = refine_profile_by_consignee(profile, consignee)
    out['Currency'] = PROFILES[profile]['currency']
    return out, pl, details, profile


def refine_profile_by_consignee(profile, consignee_lines):
    """Уточнение профиля по получателю из Details/Sheet2."""
    if profile != 'fa_logistics':
        return profile
    text = ' '.join(consignee_lines).lower()
    if 'avr' in text or 'аvr' in text:
        return 'avr'
    return 'fa_logistics'


def parse_dim_string(s):
    if not isinstance(s, str): return (None, None, None)
    nums = re.findall(r'\d+(?:[\.,]\d+)?', s)
    nums = [float(n.replace(',', '.')) for n in nums]
    return tuple(nums + [None] * (3 - len(nums)))[:3]


def read_source(src_xlsx: Path):
    xl = pd.ExcelFile(src_xlsx)
    # Формат INV+PIV HSC (тип DOC_FOR_WAY), независимо от названия файла/компании (готовый pivot) + PL
    if 'INV' in xl.sheet_names and 'PIV HSC' in xl.sheet_names:
        return _read_piv_hsc_format(src_xlsx)
    if 'TDSheet' in xl.sheet_names and 'Sheet1' not in xl.sheet_names:
        # Самодостаточный TDSheet с inline-шапкой (формат SI0245)
        return _read_tdsheet_inline_format(src_xlsx)
    if 'TDSheet' in xl.sheet_names:
        return _read_tdsheet_format(src_xlsx, xl)

    raw_inv = pd.read_excel(src_xlsx, sheet_name='Invoice List', header=0)
    raw_inv.columns = [str(c).strip() for c in raw_inv.columns]

    # Если нет колонки N — добавляем
    if 'N' not in raw_inv.columns:
        raw_inv.insert(0, 'N', range(1, len(raw_inv) + 1))

    # Фильтрация: строки с непустым Part number, исключая Total
    pn = raw_inv.get('Part number', pd.Series([None] * len(raw_inv)))
    brand = raw_inv.get('Brand', pd.Series([''] * len(raw_inv)))
    mask = pn.notna() & ~brand.astype(str).str.strip().str.lower().eq('total')
    # бэкап-фильтр по N, если Part number отсутствует
    if mask.sum() == 0:
        mask = raw_inv['N'].apply(lambda x: isinstance(x, (int, float)) and pd.notna(x))
    raw_inv = raw_inv[mask].reset_index(drop=True)
    raw_inv['N'] = range(1, len(raw_inv) + 1)

    profile = detect_profile(raw_inv.columns)
    out = pd.DataFrame()
    out['N']           = raw_inv['N'].astype(int)
    out['Brand']       = raw_inv.get('Brand', '')
    out['Part_number'] = raw_inv.get('Part number', '')
    out['Description'] = raw_inv.get('Description', '')

    desc_ru_col = 'Description Rus' if 'Description Rus' in raw_inv.columns else \
                  'Description Russian' if 'Description Russian' in raw_inv.columns else None
    out['Description_RU'] = raw_inv[desc_ru_col] if desc_ru_col else ''
    out['Quantity'] = raw_inv['Quantity'].astype(int)

    if 'Total Weight' in raw_inv.columns:
        out['Weight_unit'] = raw_inv['Weight']
        out['Gross']       = raw_inv['Total Weight']
    else:
        out['Weight_unit'] = raw_inv['Weight']
        out['Gross']       = raw_inv['Weight'] * raw_inv['Quantity']

    if profile == 'autoseller':
        out['Price']  = raw_inv['Price AED']
        out['Amount'] = raw_inv['Total AED']
    else:
        out['Price']  = raw_inv['Price USD']
        out['Amount'] = raw_inv['Total USD']

    out['Origin']   = raw_inv.get('Origin', '')
    out['HSCode']   = raw_inv['HSCode'].astype('Int64').astype(str)
    out['Currency'] = PROFILES[profile]['currency']
    out['BoxN'] = ''

    # packing list
    raw_pl = pd.read_excel(src_xlsx, sheet_name='Packing List', header=0)
    raw_pl.columns = [str(c).strip() for c in raw_pl.columns]
    pl = pd.DataFrame()
    if 'Box N' in raw_pl.columns:
        # допускаем 'PACKAGE # N' (строки) и числовые ID (190246)
        def is_box(x):
            if isinstance(x, (int, float)) and pd.notna(x): return True
            if isinstance(x, str) and re.match(r'^\s*PACKAGE\s*#', x, re.I): return True
            return False
        mask = raw_pl['Box N'].apply(is_box)
        r = raw_pl[mask].reset_index(drop=True)
        pl['Box']    = r['Box N'].apply(lambda x: f'PACKAGE # {int(x)}' if isinstance(x, (int, float)) else str(x))
        pl['Width']  = r.get('Width')
        pl['Height'] = r.get('Height')
        pl['Length'] = r.get('Length')
        pl['Weight'] = r.get('Weight')
        pl['CBM']    = r.get('CBM')
    elif 'Colli' in raw_pl.columns:
        mask = raw_pl['Colli'].apply(lambda x: isinstance(x, (int, float)) and pd.notna(x))
        r = raw_pl[mask].reset_index(drop=True)
        pl['Box'] = r['Colli'].apply(lambda x: f'PACKAGE # {int(x)}')
        dim_col = 'Dimensions cm' if 'Dimensions cm' in r.columns else 'Dimensions'
        dims = r[dim_col].apply(parse_dim_string)
        pl['Width']  = [d[0] for d in dims]
        pl['Height'] = [d[1] for d in dims]
        pl['Length'] = [d[2] for d in dims]
        pl['Weight'] = r['Weight']
        cbm_col = 'cbm' if 'cbm' in r.columns else 'CBM'
        pl['CBM']    = r[cbm_col]
    else:
        raise RuntimeError(f'Неизвестный формат Packing List: {list(raw_pl.columns)}')

    # details — пробуем разные имена листа
    xl = pd.ExcelFile(src_xlsx)
    det_sheet = None
    for cand in ['Details', 'Sheet2', 'Details ', 'Sheet1']:
        if cand in xl.sheet_names:
            det_sheet = cand; break
    if det_sheet is None:
        raise RuntimeError(f'Не найден лист с реквизитами. Листы: {xl.sheet_names}')
    raw_det = pd.read_excel(src_xlsx, sheet_name=det_sheet, header=None)
    text = '\n'.join(str(v) for v in raw_det[0].dropna())
    inv_no = _grep(r'Tax Invoice:\s*(\S+)', text)
    date   = _grep(r'Date:\s*(\S+)', text)
    contract  = _grep(r'Contract.*?:\s*(.+)', text) or ''
    incoterms = _grep(r'Incoterms:\s*(.+)', text) or ''
    consignee = _block_lines(text, r'LLC|ООО|Получатель|Autoseller|AVR')

    if not inv_no:
        lines = [str(v).strip() for v in raw_det[0].dropna()]
        if len(lines) >= 1: inv_no = lines[0]
        if len(lines) >= 2: date = lines[1]
    if not consignee:
        consignee = PROFILES[profile]['default_consignee']

    details = {'invoice_no': inv_no or 'NO_NUMBER', 'date': date or '',
               'consignee': consignee, 'contract': contract, 'incoterms': incoterms}

    profile = refine_profile_by_consignee(profile, consignee)
    out['Currency'] = PROFILES[profile]['currency']
    out['BoxN'] = ''
    return out, pl, details, profile


def _grep(pattern, text):
    m = re.search(pattern, text)
    return m.group(1).strip() if m else None


def _block_lines(text, start_pattern):
    out, started = [], False
    for ln in text.split('\n'):
        if re.search(start_pattern, ln):
            started = True
        if started:
            if re.match(r'Contract|Incoterms|Tax Invoice|Date:', ln):
                break
            out.append(ln.strip())
    return out


def enrich(inv: pd.DataFrame) -> pd.DataFrame:
    d = inv.copy()
    d[['tnved', 'group']] = d.apply(
        lambda r: classify_row(r['Description'], r['Description_RU'], r['HSCode']),
        axis=1, result_type='expand')
    return d


def aggregate(d: pd.DataFrame) -> pd.DataFrame:
    agg = d.groupby(['tnved', 'group'], sort=False).agg(
        qty=('Quantity', 'sum'), gross=('Gross', 'sum'),
        amount=('Amount', 'sum'), cnt=('N', 'count'),
    ).reset_index()
    order_idx = {g: i for i, g in enumerate(GROUP_ORDER)}
    agg['_o'] = agg.apply(lambda r: order_idx.get((r['tnved'], r['group']), 99), axis=1)
    agg = agg.sort_values('_o').drop(columns='_o').reset_index(drop=True)
    return agg


def distribute_places(agg: pd.DataFrame, total_places: int, enriched: pd.DataFrame = None) -> pd.DataFrame:
    """Распределение МЕСТ по группам. Сумма мест ВСЕГДА равна total_places
    (числу коробок в Packing List). Распределение пропорционально количеству
    штук (Quantity), минимум 1 место для каждой непустой группы."""
    out = agg.copy()
    n = len(out)
    if n == 0 or total_places == 0:
        out['places'] = 0; return out
    if total_places < n:
        out['places'] = [1 if i < total_places else 0 for i in range(n)]
        return out
    remaining = total_places - n
    if out['qty'].sum() == 0 or remaining == 0:
        out['places'] = 1; return out
    raw = out['qty'] / out['qty'].sum() * remaining
    floor = raw.apply(int); rem = raw - floor
    leftover = remaining - floor.sum()
    order = rem.sort_values(ascending=False).index.tolist()
    for i in order[:int(leftover)]:
        floor.iloc[floor.index.get_loc(i)] += 1
    out['places'] = (floor.astype(int).values + 1)
    return out


# Стили из эталона INVOICE_TRANSIT_SI0263
THIN = Side(style='thin', color='000000')
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
FONT_NAME = 'Calibri Light'
FONT_SIZE = 8
F_BOLD = Font(name=FONT_NAME, size=FONT_SIZE, bold=True)
F_REG  = Font(name=FONT_NAME, size=FONT_SIZE)
AL_C = Alignment(horizontal='center', vertical='center', wrap_text=True)
AL_L = Alignment(horizontal='left', vertical='center', wrap_text=True)
AL_R = Alignment(horizontal='right', vertical='center')

# Бухгалтерские форматы (как в эталоне)
FMT_INT    = r'_(* #,##0_);_(* \(#,##0\);_(* "-"??_);_(@_)'
FMT_MONEY  = r'_(* #,##0.00_);_(* \(#,##0.00\);_(* "-"??_);_(@_)'
FMT_WEIGHT = r'_(* #,##0.000_);_(* \(#,##0.000\);_(* "-"??_);_(@_)'
FMT_DATE   = r'[$-409]mmmm\ d\,\ yyyy;@'
ROW_H = 10.35


def _set(ws, coord, value, font=F_REG, align=AL_L, fmt=None, border=None):
    c = ws[coord]
    c.value = value; c.font = font; c.alignment = align
    if fmt: c.number_format = fmt
    if border: c.border = border


def _apply_page_setup(ws, row_count):
    """Применяет landscape, нулевые поля, единую высоту строк (как в эталоне)."""
    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    ws.page_margins.left = 0
    ws.page_margins.right = 0
    ws.page_margins.top = 0.08
    ws.page_margins.bottom = 0.31
    for r in range(1, row_count + 1):
        ws.row_dimensions[r].height = ROW_H


def build_invoice_transit(agg, details, profile, out_path: Path, blank=False, display_invoice_no=None):
    wb = Workbook(); ws = wb.active; ws.title = 'INVOICE'
    sender = [] if blank else PROFILES[profile]['sender']
    consignee = [] if blank else (details['consignee'] or PROFILES[profile]['default_consignee'])
    cur = PROFILES[profile]['currency']
    inv_no_shown = display_invoice_no or details['invoice_no']

    for i, line in enumerate(sender):
        _set(ws, f'B{3+i}', line, font=F_REG, align=AL_L)

    # Дата
    _set(ws, 'B16', 'Date:', font=F_BOLD, align=AL_L)
    date_val = details.get('date', '')
    # пытаемся разобрать дату для красивого формата
    from datetime import datetime
    parsed = None
    for fmt in ('%d.%m.%Y', '%Y-%m-%d', '%m/%d/%y', '%d/%m/%Y', '%m/%d/%Y'):
        try: parsed = datetime.strptime(str(date_val), fmt); break
        except: pass
    if parsed:
        _set(ws, 'C16', parsed, font=F_REG, align=AL_L, fmt=FMT_DATE)
    else:
        _set(ws, 'C16', date_val, font=F_REG, align=AL_L)

    # Получатель (B18:B21)
    for i, line in enumerate(consignee[:4]):
        _set(ws, f'B{18+i}', line, font=F_BOLD if i == 0 else F_REG, align=AL_L)

    # ИНВОЙС № (B22 merge B:C, D22 номер)
    _set(ws, 'B22', 'ИНВОЙС', font=F_BOLD, align=AL_L)
    ws.merge_cells('B22:C22')
    _set(ws, 'D22', inv_no_shown, font=F_BOLD, align=AL_L)

    # Заголовки таблицы (B24:H24) — все left/center, как в эталоне
    headers = ['N', 'HC CODE', 'КОЛИЧЕСТВО', 'МЕСТ', 'СУММА', 'БРУТТО', ' ОПИСАНИЕ']
    for j, h in enumerate(headers):
        _set(ws, f'{get_column_letter(2+j)}24', h, font=F_BOLD, align=AL_L, border=BOX)

    # Данные (B25..H..) — выравнивание left/center даже у чисел, формат бухгалтерский
    start = 25
    for i, row in agg.iterrows():
        r = start + i
        _set(ws, f'B{r}', i+1,                  border=BOX, align=AL_L)
        _set(ws, f'C{r}', row['tnved'],         border=BOX, align=AL_L)
        _set(ws, f'D{r}', int(row['qty']),      border=BOX, align=AL_L, fmt='General')
        _set(ws, f'E{r}', int(row['places']),   border=BOX, align=AL_L, fmt='General')
        _set(ws, f'F{r}', float(row['amount']), border=BOX, align=AL_L, fmt='0.00')
        _set(ws, f'G{r}', round(float(row['gross']), 3), border=BOX, align=AL_L, fmt='0.00')
        _set(ws, f'H{r}', row['group'],         border=BOX, align=AL_L)

    last = start + len(agg) - 1
    tr = last + 2  # пропуск строки, как в эталоне (B33..)

    _set(ws, f'B{tr}',   'КОЛИЧЕСТВО',  font=F_BOLD, align=AL_L)
    _set(ws, f'C{tr}',   f'=SUM(D{start}:D{last})', align=AL_R, fmt=FMT_INT)
    _set(ws, f'B{tr+1}', 'СУММА ДИРХАМ' if cur == 'AED' else f'СУММА {cur}', font=F_BOLD, align=AL_L)
    _set(ws, f'C{tr+1}', f'=SUM(F{start}:F{last})', align=AL_R, fmt=FMT_MONEY)
    _set(ws, f'B{tr+2}', 'БРУТТО', font=F_BOLD, align=AL_L)
    _set(ws, f'C{tr+2}', f'=SUM(G{start}:G{last})', align=AL_R, fmt=FMT_WEIGHT)
    _set(ws, f'B{tr+3}', 'МЕСТ', font=F_BOLD, align=AL_L)
    _set(ws, f'C{tr+3}', f'=SUM(E{start}:E{last})', align=AL_R, fmt=FMT_INT)

    # Ширины колонок строго как в эталоне (точные дробные значения)
    widths = {'A': 3.140625, 'B': 13.0, 'C': 14.28515625, 'D': 12.140625,
              'E': 8.85546875, 'F': 10.140625, 'G': 9.140625, 'H': 68.140625,
              'I': 3.140625, 'J': 4.85546875}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w
    _apply_page_setup(ws, tr + 5)
    wb.save(out_path)



def build_cmr_table(agg, details, profile, total_pl, out_path: Path, blank=False):
    """Таблица товаров для CMR.
    TODO (в дальнейшем): выводить CMR в формате .doc (Word), а не xlsx.
    Сейчас функция не вызывается из main — задел на будущее."""
    wb = Workbook(); ws = wb.active; ws.title = 'CMR'
    sender = [] if blank else PROFILES[profile]['sender']
    consignee = [] if blank else (details['consignee'] or PROFILES[profile]['default_consignee'])
    _set(ws, 'B2', 'CMR — данные для накладной', font=F_BOLD)
    if sender:
        _set(ws, 'B4', 'Отправитель:', font=F_BOLD)
        for i, line in enumerate(sender): _set(ws, f'C{4+i}', line)
    if consignee:
        _set(ws, 'B10', 'Получатель:', font=F_BOLD)
        for i, line in enumerate(consignee[:4]):
            _set(ws, f'C{10+i}', line, font=F_BOLD if i == 0 else F_REG)
    _set(ws, 'B16', 'Инвойс:', font=F_BOLD); _set(ws, 'C16', details['invoice_no'])
    _set(ws, 'B17', 'Дата:',   font=F_BOLD); _set(ws, 'C17', details.get('date',''))
    _set(ws, 'B18', 'Incoterms:', font=F_BOLD); _set(ws, 'C18', details.get('incoterms',''))

    headers = ['N', 'ТНВЭД', 'Наименование груза', 'Кол-во шт.', 'Мест', 'Вес брутто, кг']; hr = 22
    for j, h in enumerate(headers):
        _set(ws, f'{get_column_letter(2+j)}{hr}', h, font=F_BOLD, border=BOX, align=AL_L)

    start = hr + 1
    for i, row in agg.iterrows():
        r = start + i
        _set(ws, f'B{r}', i+1,             border=BOX, align=AL_L)
        _set(ws, f'C{r}', row['tnved'],    border=BOX, align=AL_L)
        _set(ws, f'D{r}', row['group'],    border=BOX, align=AL_L)
        _set(ws, f'E{r}', int(row['qty']), border=BOX, align=AL_L, fmt='General')
        _set(ws, f'F{r}', int(row['places']), border=BOX, align=AL_L, fmt='General')
        _set(ws, f'G{r}', round(float(row['gross']),3), border=BOX, align=AL_L, fmt='0.000')

    last = start + len(agg) - 1; tr = last + 2
    _set(ws, f'C{tr}', 'ИТОГО:', font=F_BOLD, align=AL_R)
    _set(ws, f'E{tr}', f'=SUM(E{start}:E{last})', font=F_BOLD, fmt=FMT_INT,    align=AL_R, border=BOX)
    _set(ws, f'F{tr}', f'=SUM(F{start}:F{last})', font=F_BOLD, fmt=FMT_INT,    align=AL_R, border=BOX)
    _set(ws, f'G{tr}', f'=SUM(G{start}:G{last})', font=F_BOLD, fmt=FMT_WEIGHT, align=AL_R, border=BOX)
    _set(ws, f'B{tr+2}', f'Всего мест по Packing List: {int(total_pl)}', font=F_BOLD)

    for col, w in {'A':3.14,'B':5,'C':14,'D':38,'E':10,'F':8,'G':14}.items():
        ws.column_dimensions[col].width = w
    _apply_page_setup(ws, tr + 5)
    wb.save(out_path)


def main():
    args = sys.argv[1:]
    # Скрипт создаёт ТОЛЬКО INVOICE_TRANSIT.
    # (Packing List и детальный инвойс больше не генерируются.
    #  CMR в дальнейшем будет выводиться в формате .doc — см. build_cmr_table.)
    with_header  = '--with-header' in args   # включить шапку отправителя и адрес получателя
    args = [a for a in args if not a.startswith('--')]
    if not args:
        print('Usage: python3 generate_si_documents.py <source> [out_dir] [--with-header]')
        print('  Создаёт только INVOICE_TRANSIT.')
        print('  --with-header  — включить отправителя и получателя в шапку')
        sys.exit(1)
    blank = not with_header

    src = Path(args[0]).resolve()
    out_dir = Path(args[1]).resolve() if len(args) > 1 else src.parent / 'out'
    out_dir.mkdir(parents=True, exist_ok=True)

    src_xlsx = to_xlsx_if_needed(src)
    inv_df, pl_df, details, profile = read_source(src_xlsx)

    # Если рядом лежит файл с тем же префиксом и словом 'PackingList' — подгружаем
    # из него Total Weight (брутто с упаковкой) и Total Packages (точное число коробок).
    override_gross = None
    override_places = None
    # Из details (формат INV+PIV HSC — брутто из шапки INV)
    if details.get('_override_gross'):
        override_gross = details['_override_gross']
    si_stem = re.sub(r'_Invoice.*$', '', src.stem)
    for cand in src.parent.glob(f'{si_stem}*PackingList*'):
        try:
            cand_xlsx = to_xlsx_if_needed(cand)
            pl_raw = pd.read_excel(cand_xlsx, sheet_name=0, header=None)
            for _, row in pl_raw.iterrows():
                vals = [str(v).strip() for v in row if pd.notna(v)]
                joined = ' '.join(vals)
                if joined.startswith('Total Weight'):
                    nums = [v for v in row if isinstance(v, (int, float)) and pd.notna(v)]
                    if nums: override_gross = float(nums[0])
                if 'Total Packages' in joined or 'TOTAL NUMBER OF PACKING UNITS' in joined.upper():
                    nums = [v for v in row if isinstance(v, (int, float)) and pd.notna(v)]
                    if nums: override_places = int(nums[0])
            break
        except Exception:
            pass

    # Имя выходного файла:
    si_matches = re.findall(r'SI\d+', src.stem)
    if si_matches:
        INV = '-'.join(si_matches)
    elif details['invoice_no'] and details['invoice_no'] != 'NO_NUMBER':
        INV = re.sub(r'[^A-Za-z0-9_-]', '_', details['invoice_no'])
    else:
        INV = re.sub(r'[^A-Za-z0-9_-]', '_', src.stem)

    enriched = enrich(inv_df)
    agg = aggregate(enriched)
    total_places = override_places if override_places is not None else len(pl_df)
    agg = distribute_places(agg, total_places, enriched)

    # Если есть override_gross — масштабируем брутто по группам, чтобы итог совпал
    if override_gross is not None and agg['gross'].sum() > 0:
        scale = override_gross / agg['gross'].sum()
        agg['gross'] = agg['gross'] * scale

    p_transit = out_dir / f'INVOICE_TRANSIT_{INV}.xlsx'
    build_invoice_transit(agg, details, profile, p_transit, blank=blank, display_invoice_no=INV)
    created = [p_transit]

    cur = PROFILES[profile]['currency']
    print(f'Профиль: {profile}   Валюта: {cur}')
    print(f'Инвойс: {details["invoice_no"]}    Дата: {details.get("date","")}')
    print(f'Позиций: {len(inv_df)}    Палет: {total_places}')
    print(f'Кол-во: {int(enriched["Quantity"].sum())}   '
          f'Брутто: {enriched["Gross"].sum():.2f}   '
          f'Сумма {cur}: {enriched["Amount"].sum():.2f}')
    print('\nГруппировка по ТНВЭД:')
    for _, row in agg.iterrows():
        print(f"  {row['tnved']}  {row['group']:35s} "
              f"кол-во={int(row['qty']):>5}  брутто={row['gross']:>8.2f}  "
              f"сумма={row['amount']:>10.2f}  места={int(row['places'])}")
    print(f'\nСоздано:')
    for p in created:
        print(f'  {p}')


if __name__ == '__main__':
    main()
