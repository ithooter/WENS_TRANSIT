"""
fill_inv_pl.py — заполнение шаблона INV (RU) данными из исходника.

Калибровано под шаблон OptiAuto (файл INV_OptiAuto.xlsx):
  • данные:   строки 23..185 (колонки B..S — вход; T..AB — формулы шаблона)
  • итоги:    строка 186 (G186=SUBTOTAL(9,G23:G185), L186=SUM(L23:L185))
  • шапка:    N14=Дата, N15=Номер инвойса, N18=Валюта
              BILL TO: B15.., SHIP TO: E15.. ; логотипы — картинки в шапке

Источник (Combined.xlsx, лист Sheet1): колонки определяются по заголовку.
Русское наименование берётся из колонки СРАЗУ ПОСЛЕ «Description» (без заголовка).
По стандартному правилу проекта перевод НЕ выдумывается: если русского нет —
ячейка остаётся пустой.

ВНИМАНИЕ про шапку:
  header_type = 'optiauto' → логотип из шаблона остаётся как есть.
  header_type = 'none'     → логотипы удаляются, сверху вписывается адрес отправителя.
  header_type = 'wens'     → нужен логотип/шаблон Wens (пока не реализован: ведём
                             себя как 'none' и помечаем в отчёте).
"""
from __future__ import annotations

import datetime as _dt
import shutil
from copy import copy
from pathlib import Path

import openpyxl
from openpyxl.drawing.image import Image as _XLImage

# Логотип/адрес Wens (для типа шапки 'wens')
_ASSETS = Path(__file__).resolve().parent.parent / "templates_xlsx" / "assets"
WENS_LOGO = _ASSETS / "wens_logo.png"
WENS_SENDER = [
    "DWC Business Center",
    "Level - 4, Building - A1",
    "Dubai South Business Park",
    "P.O. Box 390667",
    "Dubai, U.A.E.",
    "T: +971 4 820 8114",
]


def _set_header_logos(ws, header_type, sender_lines, sender_row0, sender_col=2,
                      logo_anchor="A1", logo_w=190, logo_h=94):
    """Управляет логотипами/адресом в шапке по типу.
      optiauto → ничего не трогаем (логотипы шаблона остаются)
      wens     → убираем логотипы шаблона и ставим логотип Wens
      none     → убираем логотипы и вписываем адрес отправителя
    Возвращает список заметок.
    """
    notes = []
    if header_type == "optiauto":
        return notes
    if hasattr(ws, "_images"):
        ws._images = []
    if header_type == "wens":
        if WENS_LOGO.exists():
            img = _XLImage(str(WENS_LOGO))
            img.width, img.height = logo_w, logo_h
            ws.add_image(img, logo_anchor)
        else:
            notes.append("логотип Wens не найден в assets/")
        # адрес отправителя: из формы, иначе фирменный адрес Wens
        for k, line in enumerate((sender_lines or WENS_SENDER)[:6]):
            ws.cell(row=sender_row0 + k, column=sender_col).value = line
    elif header_type == "none":
        for k, line in enumerate((sender_lines or [])[:6]):
            ws.cell(row=sender_row0 + k, column=sender_col).value = line
    return notes

try:
    from hssc_txt_v3 import country_to_iso2
except Exception:  # на случай отдельного запуска
    def country_to_iso2(name):  # noqa
        return ""

# --- геометрия шаблона ---
DATA_START = 23
DATA_END = 185
TOTAL_ROW = 186            # TOTAL AMOUNT (SUBTOTAL/SUM по данным)
HDR_DATE = (14, 14)        # N14
HDR_INVNO = (15, 14)       # N15
HDR_CURR = (18, 14)        # N18
BILL_TO_ROW0 = 15          # B15..  (получатель)
SHIP_TO_COL = 5            # E15..  (зеркало)
SENDER_ROW0 = 2            # B2.. (для режима «без шапки»)

# логические поля → колонка шаблона INV (1-based)
COL = {
    "n": 2, "brand": 3, "part": 4, "desc_ru": 5, "qty": 7, "weight": 8,
    "price": 9, "amt": 12, "vat_rate": 14, "vat_amt": 15, "coo": 16,
    "hs": 17, "alpha2": 19,
}
HELPER_COLS = list(range(20, 31))  # T..AD — формулы-справочники шаблона


def _norm(s):
    return str(s).strip().lower() if s is not None else ""


def _parse_date(d):
    if d in (None, ""):
        return None
    if isinstance(d, (_dt.datetime, _dt.date)):
        return d
    for fmt in ("%d.%m.%Y", "%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y", "%d %B %Y", "%d %b %Y"):
        try:
            return _dt.datetime.strptime(str(d).strip(), fmt)
        except ValueError:
            continue
    return str(d)


def _detect_source_cols(ws):
    """Ищет строку заголовка и возвращает (map, header_row, currency)."""
    for r in range(1, 12):
        labels = {_norm(ws.cell(row=r, column=c).value): c
                  for c in range(1, ws.max_column + 1)}
        if "brand" in labels and any(k in labels for k in ("part number", "part no")):
            m = {}
            m["brand"] = labels.get("brand")
            m["part"] = labels.get("part number") or labels.get("part no")
            desc_c = labels.get("description")
            m["desc_en"] = desc_c
            m["desc_ru"] = (desc_c + 1) if desc_c else None  # колонка RU без заголовка
            m["weight"] = labels.get("weight")
            m["qty"] = labels.get("quantity") or labels.get("qty")
            m["price"] = labels.get("price aed") or labels.get("price usd") or labels.get("price")
            m["amt"] = (labels.get("total aed") or labels.get("total usd")
                        or labels.get("amount") or labels.get("total"))
            m["coo"] = labels.get("origin") or labels.get("coo") or labels.get("country")
            m["hs"] = labels.get("hs code") or labels.get("hscode") or labels.get("hs")
            currency = "AED" if "total aed" in labels or "price aed" in labels else \
                       "USD" if "total usd" in labels or "price usd" in labels else "AED"
            return m, r, currency
    raise RuntimeError("В исходнике не найден заголовок таблицы (Brand / Part Number).")


_TOTAL_PREFIXES = ("total", "итого", "всего", "grand", "subtotal", "итог")


def _read_rows(ws, m, header_row):
    rows = []
    for r in range(header_row + 1, ws.max_row + 1):
        brand = ws.cell(row=r, column=m["brand"]).value if m["brand"] else None
        part = ws.cell(row=r, column=m["part"]).value if m["part"] else None
        if (brand in (None, "")) and (part in (None, "")):
            continue
        # отсеиваем итоговые строки источника («Total Amount» и т.п.)
        btxt = str(brand).strip().lower() if brand else ""
        ptxt = str(part).strip().lower() if part else ""
        if btxt.startswith(_TOTAL_PREFIXES) or ptxt.startswith(_TOTAL_PREFIXES):
            continue

        def g(key):
            c = m.get(key)
            return ws.cell(row=r, column=c).value if c else None

        rows.append({
            "brand": brand, "part": part,
            "desc_ru": g("desc_ru"), "desc_en": g("desc_en"),
            "weight": g("weight"), "qty": g("qty"),
            "price": g("price"), "amt": g("amt"),
            "coo": g("coo"), "hs": g("hs"),
        })
    return rows


def fill_inv_ru(*, template_path, source_path, output_path,
                invoice_no=None, date=None, currency=None,
                receiver_lines=None, sender_lines=None,
                header_type="optiauto", source_sheet="Sheet1"):
    """Заполняет INV(RU) из исходника. Возвращает отчёт-словарь."""
    receiver_lines = receiver_lines or []
    sender_lines = sender_lines or []
    notes = []

    swb = openpyxl.load_workbook(source_path, data_only=True)
    sws = swb[source_sheet] if source_sheet in swb.sheetnames else swb[swb.sheetnames[0]]
    m, hr, src_currency = _detect_source_cols(sws)
    rows = _read_rows(sws, m, hr)
    swb.close()

    shutil.copy(template_path, output_path)
    wb = openpyxl.load_workbook(output_path)
    ws = wb["INV"]

    # --- при необходимости расширяем ёмкость (если строк больше 163) ---
    n = len(rows)
    capacity = DATA_END - DATA_START + 1
    total_row = TOTAL_ROW
    if n > capacity:
        extra = n - capacity
        ws.insert_rows(DATA_END + 1, extra)
        for r in range(DATA_END + 1, DATA_END + extra + 1):
            for c in range(1, 31):
                src = ws.cell(row=DATA_START, column=c)
                dst = ws.cell(row=r, column=c)
                dst.font = copy(src.font); dst.border = copy(src.border)
                dst.fill = copy(src.fill); dst.alignment = copy(src.alignment)
                dst.number_format = src.number_format
            for c in HELPER_COLS:  # перенос формул-справочников с заменой номера строки
                f = ws.cell(row=DATA_START, column=c).value
                if isinstance(f, str) and f.startswith("="):
                    ws.cell(row=r, column=c).value = f.replace(str(DATA_START), str(r))
        total_row = TOTAL_ROW + extra
        end = DATA_START + n - 1
        ws.cell(row=total_row, column=7).value = f"=SUBTOTAL(9,G{DATA_START}:G{end})"
        ws.cell(row=total_row, column=12).value = f"=SUM(L{DATA_START}:L{end})"
        ws.cell(row=total_row, column=14).value = f"=SUM(N{DATA_START}:N{end})"
        ws.cell(row=total_row + 2, column=12).value = f"=L{total_row}"
        ws.cell(row=16, column=14).value = f"=G{total_row}"
        ws.cell(row=17, column=14).value = f"=L{total_row + 2}"
        notes.append(f"Строк больше шаблона ({n}>{capacity}) — вставлено {extra}; проверь итоги.")

    # --- данные ---
    for i, rec in enumerate(rows):
        r = DATA_START + i
        qty = rec["qty"] or 0
        price = rec["price"]
        amt = rec["amt"] if rec["amt"] not in (None, "") else \
            (qty * price if isinstance(price, (int, float)) else None)
        ws.cell(row=r, column=COL["n"]).value = i + 1
        ws.cell(row=r, column=COL["brand"]).value = rec["brand"]
        ws.cell(row=r, column=COL["part"]).value = rec["part"]
        # RU-наименование: если пусто — НЕ выдумываем перевод, оставляем пусто
        ws.cell(row=r, column=COL["desc_ru"]).value = rec["desc_ru"] or None
        ws.cell(row=r, column=COL["qty"]).value = qty
        ws.cell(row=r, column=COL["weight"]).value = rec["weight"]
        ws.cell(row=r, column=COL["price"]).value = price
        ws.cell(row=r, column=COL["amt"]).value = amt
        ws.cell(row=r, column=COL["vat_rate"]).value = 0
        ws.cell(row=r, column=COL["vat_amt"]).value = 0
        ws.cell(row=r, column=COL["coo"]).value = rec["coo"]
        ws.cell(row=r, column=COL["hs"]).value = rec["hs"]
        ws.cell(row=r, column=COL["alpha2"]).value = country_to_iso2(rec["coo"]) or None

    # --- чистим неиспользованные строки данных (вместе с формулами-справочниками) ---
    for r in range(DATA_START + n, DATA_END + 1):
        for c in range(1, 31):
            ws.cell(row=r, column=c).value = None

    # --- шапка: дата / номер / валюта ---
    pdate = _parse_date(date)
    if pdate is not None:
        ws.cell(row=HDR_DATE[0], column=HDR_DATE[1]).value = pdate
    if invoice_no:
        ws.cell(row=HDR_INVNO[0], column=HDR_INVNO[1]).value = invoice_no
    ws.cell(row=HDR_CURR[0], column=HDR_CURR[1]).value = currency or src_currency

    # --- получатель (BILL TO + SHIP TO) ---
    if receiver_lines:
        for k in range(6):
            r = BILL_TO_ROW0 + k
            val = receiver_lines[k] if k < len(receiver_lines) else None
            ws.cell(row=r, column=2).value = val
            ws.cell(row=r, column=SHIP_TO_COL).value = val

    # --- тип шапки (логотипы / адрес отправителя) ---
    notes += _set_header_logos(ws, header_type, sender_lines, SENDER_ROW0, sender_col=2)

    wb.save(output_path)
    wb.close()
    return {
        "output": str(output_path),
        "rows": n,
        "currency": currency or src_currency,
        "header_type": header_type,
        "notes": notes,
    }


# --------------------------------------------------------------------------- #
# PL — упаковочный лист
# --------------------------------------------------------------------------- #

# геометрия PL-шаблона
PL_DATA_START = 21
PL_DATA_END = 41          # вместимость 21 коробка; итоги 43..45
PL_TOTAL_PKG = 43         # C43=SUM(F21:F41)
PL_DATE = (10, 2)         # B10
PL_INVNO = (11, 2)        # B11
PL_BILL_ROW0 = 13         # B13..  (получатель)
PL_SHIP_COL = 7           # G13..  (зеркало)
PL_SENDER_ROW0 = 2        # для «без шапки»

PL_COL = {"pkg": 2, "w": 3, "h": 4, "l": 5, "qty": 6, "weight": 7, "cbm": 8}


def _detect_pl_cols(ws):
    """Находит строку заголовка PL (есть 'Box N' и 'Width') и колонки."""
    for r in range(1, 12):
        labels = {_norm(ws.cell(row=r, column=c).value): c
                  for c in range(1, ws.max_column + 1)}
        if any(k in labels for k in ("box n", "box number", "номер коробки")) \
                and ("width" in labels or "ширина" in labels):
            m = {
                "box": labels.get("box n") or labels.get("box number") or labels.get("номер коробки"),
                "w": labels.get("width") or labels.get("ширина"),
                "h": labels.get("height") or labels.get("высота"),
                "l": labels.get("length") or labels.get("длина"),
                "weight": labels.get("weight") or labels.get("вес"),
                "cbm": labels.get("cbm") or labels.get("объем"),
                "qty": labels.get("qty") or labels.get("quantity") or labels.get("кол-во мест"),
            }
            return m, r
    raise RuntimeError("В исходнике не найден заголовок упаковки (Box N / Width).")


def _read_pl_rows(ws, m, header_row):
    rows = []
    for r in range(header_row + 1, ws.max_row + 1):
        box = ws.cell(row=r, column=m["box"]).value if m["box"] else None
        if box in (None, ""):
            continue
        if str(box).strip().lower().startswith(_TOTAL_PREFIXES):
            continue

        def g(key):
            c = m.get(key)
            return ws.cell(row=r, column=c).value if c else None
        rows.append({"box": box, "w": g("w"), "h": g("h"), "l": g("l"),
                     "weight": g("weight"), "cbm": g("cbm"), "qty": g("qty")})
    return rows


def fill_pl(*, template_path, source_path, output_path,
            invoice_no=None, date=None, receiver_lines=None, sender_lines=None,
            header_type="optiauto", source_sheet="Sheet2"):
    """Заполняет PL (упаковочный лист) из исходника. Возвращает отчёт."""
    receiver_lines = receiver_lines or []
    sender_lines = sender_lines or []
    notes = []

    swb = openpyxl.load_workbook(source_path, data_only=True)
    sws = swb[source_sheet] if source_sheet in swb.sheetnames else swb[swb.sheetnames[-1]]
    m, hr = _detect_pl_cols(sws)
    rows = _read_pl_rows(sws, m, hr)
    swb.close()

    shutil.copy(template_path, output_path)
    wb = openpyxl.load_workbook(output_path)
    ws = wb["PL"]

    n = len(rows)
    capacity = PL_DATA_END - PL_DATA_START + 1
    total_pkg_row = PL_TOTAL_PKG
    if n > capacity:
        extra = n - capacity
        ws.insert_rows(PL_DATA_END + 1, extra)
        for r in range(PL_DATA_END + 1, PL_DATA_END + extra + 1):
            for c in range(1, 12):
                src = ws.cell(row=PL_DATA_START, column=c)
                dst = ws.cell(row=r, column=c)
                dst.font = copy(src.font); dst.border = copy(src.border)
                dst.fill = copy(src.fill); dst.alignment = copy(src.alignment)
                dst.number_format = src.number_format
        total_pkg_row = PL_TOTAL_PKG + extra
        end = PL_DATA_START + n - 1
        ws.cell(row=total_pkg_row, column=3).value = f"=SUM(F{PL_DATA_START}:F{end})"
        ws.cell(row=total_pkg_row + 1, column=3).value = f"=SUM(G{PL_DATA_START}:G{end})"
        ws.cell(row=total_pkg_row + 2, column=3).value = f"=SUM(H{PL_DATA_START}:H{end})"
        notes.append(f"Коробок больше шаблона ({n}>{capacity}) — вставлено {extra}; проверь итоги.")

    for i, rec in enumerate(rows):
        r = PL_DATA_START + i
        ws.cell(row=r, column=PL_COL["pkg"]).value = i + 1
        ws.cell(row=r, column=PL_COL["w"]).value = rec["w"]
        ws.cell(row=r, column=PL_COL["h"]).value = rec["h"]
        ws.cell(row=r, column=PL_COL["l"]).value = rec["l"]
        ws.cell(row=r, column=PL_COL["qty"]).value = rec["qty"] if rec["qty"] not in (None, "") else 1
        ws.cell(row=r, column=PL_COL["weight"]).value = rec["weight"]
        ws.cell(row=r, column=PL_COL["cbm"]).value = rec["cbm"]

    for r in range(PL_DATA_START + n, PL_DATA_END + 1):
        for c in range(2, 12):
            ws.cell(row=r, column=c).value = None

    pdate = _parse_date(date)
    if pdate is not None:
        ws.cell(row=PL_DATE[0], column=PL_DATE[1]).value = pdate
    if invoice_no:
        ws.cell(row=PL_INVNO[0], column=PL_INVNO[1]).value = invoice_no

    if receiver_lines:
        for k in range(6):
            r = PL_BILL_ROW0 + k
            val = receiver_lines[k] if k < len(receiver_lines) else None
            ws.cell(row=r, column=2).value = val
            ws.cell(row=r, column=PL_SHIP_COL).value = val

    notes += _set_header_logos(ws, header_type, sender_lines, PL_SENDER_ROW0, sender_col=1)

    wb.save(output_path)
    wb.close()
    return {"output": str(output_path), "rows": n, "header_type": header_type, "notes": notes}
