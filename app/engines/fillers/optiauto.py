"""
optiauto.py — заполнитель шаблона OptiAuto (единый workbook, листы INV + PL).

Читает canonical.Shipment и заполняет ОБА листа за один проход через общий
sheet_writer (одна механика на INV и PL — без дублирования). Результат — «Весь
шаблон» (заполненный workbook); INV(RU) и PL получаются извлечением нужного
листа (это делает service.py).

Геометрия — под реальные файлы templates_xlsx/OptiAuto/*.xlsx:
  INV: заголовки строка 22, данные 23..185, итоги 186/187/188, шапка N14/N15/N18,
       Num Pckg's/Vol/Weight в Q18/Q19/Q20, справочные VLOOKUP-колонки T,U,V,W,Z,AB.
  PL : заголовки 19-20, данные 21..41, итоги C43/C44/C45, помощник N=L-G.
"""
from __future__ import annotations

import shutil
from pathlib import Path

import openpyxl
from openpyxl.drawing.image import Image as XLImage
from openpyxl.worksheet.properties import PageSetupProperties

from ..canonical import Shipment, _num
from ..sheet_writer import SheetGeometry, fill_sheet

# Логотип OptiAuto и картинка-адрес (шапка). openpyxl не читает исходные
# twoCellAnchor-рисунки шаблона и теряет их при сохранении, поэтому мы
# переклеиваем их сами через add_image по точным позициям из drawing1.xml.
_ASSETS = Path(__file__).resolve().parent.parent.parent / "templates_xlsx" / "assets"
_LOGO = _ASSETS / "optiauto_logo.png"
_ADDRESS = _ASSETS / "optiauto_address.png"
# (anchor, ширина×высота в px) для логотипа и адреса на каждом листе.
_INV_IMAGES = [(_LOGO, "A1", (190, 165)), (_ADDRESS, "P2", (238, 92))]
_PL_IMAGES = [(_LOGO, "A1", (167, 149)), (_ADDRESS, "G2", (281, 114))]


def _place_images(ws, specs) -> None:
    for path, anchor, (w, h) in specs:
        if path.exists():
            img = XLImage(str(path))
            img.width, img.height = w, h
            ws.add_image(img, anchor)

try:
    from ..fill_inv_pl import _parse_date, country_to_iso2
except Exception:  # автономный запуск / отсутствует hssc
    def _parse_date(d):
        return d

    def country_to_iso2(name):
        return ""


# --------------------------------------------------------------------------- #
# Геометрия листов
# --------------------------------------------------------------------------- #

def _inv_on_expand(ws, s: int, e: int, extra: int) -> None:
    """Переписать блок ИТОГО и ссылки шапки INV после вставки строк."""
    t1, t2, t3 = 186 + extra, 187 + extra, 188 + extra
    ws.cell(row=t1, column=7).value = f"=SUBTOTAL(9,G{s}:G{e})"   # QTY
    ws.cell(row=t1, column=11).value = f"=SUM(K{s}:K{e})"          # AMT B
    ws.cell(row=t1, column=12).value = f"=SUM(L{s}:L{e})"          # AMT AED
    ws.cell(row=t1, column=13).value = f"=SUM(M{s}:M{e})"
    ws.cell(row=t1, column=14).value = f"=SUM(N{s}:N{e})"          # VAT RATE col
    ws.cell(row=t2, column=15).value = f"=SUM(O{s}:O{e})"          # VAT AMT
    ws.cell(row=t3, column=11).value = f"=K{t1}"
    ws.cell(row=t3, column=12).value = f"=L{t1}"
    ws.cell(row=t3, column=13).value = f"=M{t1}+O{t2}"
    ws.cell(row=16, column=14).value = f"=G{t1}"                   # N16 Total Units
    ws.cell(row=17, column=14).value = f"=L{t3}"                   # N17 Total Amount


def _pl_on_expand(ws, s: int, e: int, extra: int) -> None:
    """Переписать итоги PL после вставки строк."""
    ws.cell(row=43 + extra, column=3).value = f"=SUM(F{s}:F{e})"   # Total PKG's
    ws.cell(row=44 + extra, column=3).value = f"=SUM(G{s}:G{e})"   # Total Weight
    ws.cell(row=45 + extra, column=3).value = f"=SUM(H{s}:H{e})"   # Total Volume


INV_GEOM = SheetGeometry(
    sheet_name="INV",
    data_start=23,
    data_end=185,
    template_cols={
        "n": 2, "brand": 3, "part_number": 4, "description_ru": 5,
        "qty": 7, "unit_weight": 8, "unit_price": 9, "amount": 12,
        "vat_rate": 14, "vat_amt": 15, "coo": 16, "hs_code": 17, "alpha2": 19,
    },
    style_cols=(1, 31),
    helper_formula_cols=(20, 21, 22, 23, 26, 28),   # T,U,V,W,Z,AB — VLOOKUP-справочники
    renumber_field="n",
    on_expand=_inv_on_expand,
)

PL_GEOM = SheetGeometry(
    sheet_name="PL",
    data_start=21,
    data_end=41,
    template_cols={"pkg": 2, "w": 3, "h": 4, "l": 5, "qty": 6,
                   "weight": 7, "cbm": 8},
    style_cols=(1, 14),
    helper_formula_cols=(14,),                       # N = L - G
    renumber_field="pkg",
    on_expand=_pl_on_expand,
)


# --------------------------------------------------------------------------- #
# canonical → строки шаблона
# --------------------------------------------------------------------------- #

def _inv_rows(goods) -> list[dict]:
    rows = []
    for g in goods:
        rows.append({
            "brand": g.brand,
            "part_number": g.part_number,
            # Русское наименование НЕ выдумываем: нет — оставляем пустым.
            "description_ru": g.description_ru or None,
            "qty": _num(g.qty) or 0,
            "unit_weight": _num(g.unit_weight),
            "unit_price": _num(g.unit_price),
            "amount": g.effective_amount,
            "vat_rate": 0,
            "vat_amt": 0,
            "coo": g.coo,
            "hs_code": g.hs_code,
            "alpha2": country_to_iso2(g.coo) or None,
        })
    return rows


def _pl_rows(boxes) -> list[dict]:
    rows = []
    for b in boxes:
        rows.append({
            "w": _num(b.width),
            "h": _num(b.height),
            "l": _num(b.length),
            "qty": _num(b.qty) if b.qty not in (None, "") else 1,
            "weight": _num(b.gross_weight),
            "cbm": _num(b.cbm),
        })
    return rows


# --------------------------------------------------------------------------- #
# Шапка
# --------------------------------------------------------------------------- #

def _write_header(inv_ws, pl_ws, sh: Shipment, invoice_no, date, remove_logo) -> None:
    req = sh.requisites
    # INV: дата / номер / валюта
    pdate = _parse_date(date or req.date)
    if pdate not in (None, ""):
        inv_ws.cell(row=14, column=14).value = pdate               # N14
    if invoice_no or req.invoice_no:
        inv_ws.cell(row=15, column=14).value = invoice_no or req.invoice_no  # N15
    if req.currency:
        inv_ws.cell(row=18, column=14).value = req.currency        # N18
        # Заголовок колонки суммы по валюте: " AMT USD" / " AMT AED " (шаблон — AED)
        inv_ws.cell(row=22, column=12).value = f" AMT {req.currency}"  # L22

    # INV: реальные итоги мест / объёма / веса из canonical (были статичные placeholder)
    inv_ws.cell(row=18, column=17).value = sh.total_places or None  # Q18 Num Pckg's
    cbm = sum(_num(b.cbm) or 0 for b in sh.boxes)
    inv_ws.cell(row=19, column=17).value = round(cbm, 3) or None    # Q19 Vol CBM
    inv_ws.cell(row=20, column=17).value = round(sh.total_gross, 3) or None  # Q20 Weight

    # Получатель: BILL TO (col B) + SHIP TO (col E) в строках 15-20
    rlines = req.receiver_lines or []
    if rlines:
        for k in range(6):
            r = 15 + k
            val = rlines[k] if k < len(rlines) else None
            inv_ws.cell(row=r, column=2).value = val
            inv_ws.cell(row=r, column=5).value = val

    # Отправитель: только Базовый (remove_logo) пишет адрес поверх шапки (строки 2-7).
    if remove_logo:
        if hasattr(inv_ws, "_images"):
            inv_ws._images = []
        if hasattr(pl_ws, "_images"):
            pl_ws._images = []
        slines = req.sender_lines or []
        for k in range(6):
            inv_ws.cell(row=2 + k, column=2).value = (
                slines[k] if k < len(slines) else None)

    # PL: дата / номер
    if pdate not in (None, ""):
        pl_ws.cell(row=10, column=2).value = pdate                 # B10
    if invoice_no or req.invoice_no:
        pl_ws.cell(row=11, column=2).value = invoice_no or req.invoice_no  # B11
    # Получатель PL: имя в строке 13 (рядом с меткой BILL TO:), только по числу
    # реальных строк — чтобы не затирать заголовки колонок в строке 19
    # (B19="Номер Коробки", G19="Вес").
    for k, val in enumerate(rlines[:6]):
        pl_ws.cell(row=13 + k, column=2).value = val   # BILL TO
        pl_ws.cell(row=13 + k, column=7).value = val   # SHIP TO


# --------------------------------------------------------------------------- #
# Точка входа
# --------------------------------------------------------------------------- #

def fill_optiauto_whole(shipment: Shipment, template_path, output_path, *,
                        invoice_no=None, date=None, remove_logo=False) -> dict:
    """Заполнить весь OptiAuto-workbook (INV + PL) из canonical. Возвращает отчёт.

    Файл сохраняется как «Весь шаблон». INV(RU)/PL извлекаются из него отдельно.
    """
    shutil.copy(template_path, output_path)
    wb = openpyxl.load_workbook(output_path)
    inv_ws, pl_ws = wb["INV"], wb["PL"]

    inv_report = fill_sheet(inv_ws, _inv_rows(shipment.goods), INV_GEOM)
    pl_report = fill_sheet(pl_ws, _pl_rows(shipment.boxes), PL_GEOM)
    _write_header(inv_ws, pl_ws, shipment, invoice_no, date, remove_logo)

    # Логотип + картинка-адрес OptiAuto (только когда шапка сохраняется).
    if not remove_logo:
        _place_images(inv_ws, _INV_IMAGES)
        _place_images(pl_ws, _PL_IMAGES)

    # INV — вид «как на бумаге»:
    #  • прячем сетку (шапка без разделителей; у таблицы свои границы);
    #  • freeze по строку-заголовок таблицы (22) — при прокрутке видна «шапка» столбцов;
    #  • режим Page Break Preview + fit-to-page (как в шаблоне/PL) — иначе синяя
    #    рамка print_area не видна: cell-copy сборкой эти настройки терялись.
    inv_ws.sheet_view.showGridLines = False
    inv_ws.freeze_panes = "A23"
    inv_ws.sheet_view.view = "pageBreakPreview"
    inv_ws.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)
    inv_ws.page_setup.fitToWidth = 1
    inv_ws.page_setup.fitToHeight = 0

    # Динамическая рамка (print_area) — ровно по конец таблицы (последняя строка ИТОГО):
    #   INV: totals = new_end + 3;  PL: totals = new_end + 4 (после пустой строки 42).
    inv_ws.print_area = f"A1:T{inv_report['new_end'] + 3}"
    pl_ws.print_area = f"A1:I{pl_report['new_end'] + 4}"

    wb.save(output_path)
    wb.close()

    notes = []
    if inv_report["expanded"]:
        notes.append(f"INV: строк больше вместимости шаблона "
                     f"({inv_report['rows_written']}) — вставлены строки, проверь итоги.")
    if pl_report["expanded"]:
        notes.append(f"PL: коробок больше вместимости шаблона "
                     f"({pl_report['rows_written']}) — вставлены строки, проверь итоги.")
    return {
        "output": str(output_path),
        "inv_rows": inv_report["rows_written"],
        "pl_rows": pl_report["rows_written"],
        "notes": notes,
    }
