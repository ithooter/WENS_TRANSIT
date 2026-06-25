#!/usr/bin/env python3
"""
generate_customs_txt.py — Dubai customs TXT generator with company classification.

USAGE
-----
    python3 generate_customs_txt.py --Wens     --input <source.xlsx> [options]
    python3 generate_customs_txt.py --OptiAuto --input <source.xlsx> [options]

The company flag selects a profile (company name, currency, supplier code,
incoterm) from company_profiles.py. Everything the profile can't know
(declaration number, pages, exact metadata) is auto-extracted from the source
workbook where possible, or supplied via options / left for the operator.

EXAMPLES
--------
    # WENS / TOP PARTS supplier file with a ready HS CODE SUM sheet:
    python3 generate_customs_txt.py --Wens \\
        --input INV___HS_CODE_SUM_TPL2605012.xlsx \\
        --decl-no TPL2605012 --pages 1 \\
        --company "TOP PARTS GENERAL TRADING LOGISTICS" --supplier-code 7 \\
        --output TPL2605012.txt

    # OptiAuto base template:
    python3 generate_customs_txt.py --OptiAuto \\
        --input OPTIAUTO_APMG_00_2025.xlsx \\
        --output optiauto.txt

    # List registered companies:
    python3 generate_customs_txt.py --list

OPTIONS
-------
    --input PATH         source workbook (.xlsx; convert .xlsb/.xls first)
    --output PATH        output TXT path (default: <input_stem>.txt)
    --hssc-sheet NAME    sheet to read aggregated data from
                         (default: auto — 'HS CODE SUM' or 'HSSC')
    --decl-no STR        declaration number (default: auto from source)
    --decl-date STR      YYYY-MM-DD or YY-MM-DD (default: auto from source)
    --pages N            IH header page count (default: heuristic from INV)
    --company STR        override IH company name
    --supplier-code N    override IH supplier code [6]
    --total FLOAT        override total amount (default: sum of amounts)
    --incoterm STR       override incoterm (default: profile default)
    --fill-wens PATH     also fill a WENS/OptiAuto template copy at PATH
                         (writes HSSC + PIV HSC sheets)
"""

from __future__ import annotations
import argparse
import datetime as _dt
import math
import sys
from pathlib import Path

import openpyxl

import hssc_txt_v3 as engine
from company_profiles import resolve_profile, list_companies, PROFILES


# --------------------------------------------------------------------------- #
# Source metadata extraction
# --------------------------------------------------------------------------- #

def _scan_inv_metadata(wb) -> dict:
    """Pull decl_no, date, currency, total from an INV/INVOICE sheet.

    Works across template variants by searching for label cells and reading
    the value cell to their right (1–2 columns over).
    """
    meta: dict = {}
    sheet_name = None
    for cand in ("INV", "INVOICE"):
        if cand in wb.sheetnames:
            sheet_name = cand
            break
    if sheet_name is None:
        return meta
    ws = wb[sheet_name]

    label_targets = {
        "tax invoice number": "decl_no",
        "no.": "decl_no",
        "date": "date",
        "total amount": "total",
        "currency type": "currency",
    }

    for r in range(1, 30):
        for c in range(1, 25):
            v = ws.cell(row=r, column=c).value
            if not isinstance(v, str):
                continue
            key = v.strip().lower().rstrip(":")
            if key in label_targets:
                # value is usually 1–2 columns to the right
                for dc in (1, 2):
                    val = ws.cell(row=r, column=c + dc).value
                    if val not in (None, ""):
                        meta[label_targets[key]] = val
                        break
    return meta


def _fmt_date(value) -> str | None:
    """Normalize a date value (datetime or string) to YYYY-MM-DD."""
    if value is None:
        return None
    if isinstance(value, _dt.datetime):
        return value.strftime("%Y-%m-%d")
    if isinstance(value, str):
        s = value.strip()
        # try a few common formats
        for fmt in ("%d %b %Y", "%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y", "%d.%m.%Y"):
            try:
                return _dt.datetime.strptime(s, fmt).strftime("%Y-%m-%d")
            except ValueError:
                continue
        return s  # pass through, engine will handle YY conversion
    return None


def _pick_data_sheet(wb, override: str | None) -> str:
    if override:
        return override
    for cand in ("HS CODE SUM", "HSSC"):
        if cand in wb.sheetnames:
            return cand
    raise RuntimeError(
        f"No 'HS CODE SUM' or 'HSSC' sheet found. Sheets: {wb.sheetnames}"
    )


# --------------------------------------------------------------------------- #
# Main
# --------------------------------------------------------------------------- #

def build_arg_parser() -> argparse.ArgumentParser:
    p = argparse.ArgumentParser(
        prog="generate_customs_txt.py",
        description="Dubai customs TXT generator with per-company profiles.",
        formatter_class=argparse.RawDescriptionHelpFormatter,
    )
    # Company selector flags — one boolean per profile, e.g. --Wens / --OptiAuto
    grp = p.add_mutually_exclusive_group()
    for key, prof in PROFILES.items():
        primary = prof.cli_flags[0]
        grp.add_argument(
            primary, *prof.cli_flags[1:],
            dest="company_flag", action="store_const", const=key,
            help=f"Use {prof.company_name} profile ({prof.currency})",
        )
    p.add_argument("--company-key", dest="company_key",
                   help="Alternative to the company flag: 'wens' / 'optiauto'")

    p.add_argument("--input", help="Source workbook (.xlsx)")
    p.add_argument("--template", help="Template workbook to fill (fill_template mode)")
    p.add_argument("--source-inv", dest="source_inv",
                   help="Source INV book (fill_template mode)")
    p.add_argument("--source-pl", dest="source_pl",
                   help="Source PL book (fill_template mode)")
    p.add_argument("--output", help="Output path")
    p.add_argument("--hssc-sheet", dest="hssc_sheet",
                   help="Sheet with aggregated data (default: auto)")
    p.add_argument("--decl-no", dest="decl_no")
    p.add_argument("--decl-date", dest="decl_date")
    p.add_argument("--pages", type=int)
    p.add_argument("--company", dest="company_name_override")
    p.add_argument("--supplier-code", dest="supplier_code", type=int)
    p.add_argument("--total", dest="total", type=float)
    p.add_argument("--incoterm")
    p.add_argument("--cd-blank", dest="cd_blank", default="CD_BLANK_NEW.xlsx",
                   help="Path to CD_BLANK_NEW.xlsx (dt_hscodes lookup)")
    p.add_argument("--fill-wens", dest="fill_wens",
                   help="Also write HSSC+PIV HSC into a copy at this path")
    p.add_argument("--list", action="store_true",
                   help="List registered companies and exit")
    return p


def _run_fill_template(args, profile) -> int:
    """OptiAuto-style: fill INV + PL of a template, preserving design."""
    import fill_optiauto

    template = args.template
    source_inv = args.source_inv
    source_pl = args.source_pl

    missing = [n for n, v in (("--template", template),
                              ("--source-inv", source_inv),
                              ("--source-pl", source_pl)) if not v]
    if missing:
        print(f"ERROR ({profile.key} fill mode): need {', '.join(missing)}.\n"
              f"Example:\n  python3 generate_customs_txt.py --OptiAuto \\\n"
              f"    --template OPTIAUTO_APMG_00_2025.xlsx \\\n"
              f"    --source-inv INV_book.xlsx --source-pl PL_book.xlsx \\\n"
              f"    --output OPTIAUTO_filled.xlsx", file=sys.stderr)
        return 2

    out = args.output or "OPTIAUTO_filled.xlsx"
    try:
        result = fill_optiauto.fill_optiauto(
            template_path=template,
            source_inv_path=source_inv,
            source_pl_path=source_pl,
            output_path=out,
        )
    except Exception as e:
        print(f"ERROR during fill: {e}", file=sys.stderr)
        return 1

    print(f"Company:   {profile.company_name}  [{profile.key}]  (fill_template)")
    print(f"Template:  {Path(template).name}")
    print(f"Output:    {result['output']}")
    print(f"INV rows:  {result['inv_rows']}  (header detected row {result['inv_header_row']})")
    print(f"PL rows:   {result['pl_rows']}  (header detected row {result['pl_header_row']})")
    print(f"\nINV columns mapped: {result['inv_columns_detected']}")
    print(f"PL columns mapped:  {result['pl_columns_detected']}")
    print("\nDesign preserved: header, addresses, red headers (FFC00000), "
          "totals formulas. Currency read from source.")
    return 0


def main(argv=None) -> int:
    args = build_arg_parser().parse_args(argv)

    if args.list:
        print("Registered companies:")
        print(list_companies())
        return 0

    # Resolve company profile
    company_sel = args.company_flag or args.company_key
    if not company_sel:
        print("ERROR: choose a company, e.g. --Wens or --OptiAuto "
              "(see --list).", file=sys.stderr)
        return 2
    profile = resolve_profile(company_sel)

    # ── Branch by output mode ──
    if profile.mode == "fill_template":
        return _run_fill_template(args, profile)

    if not args.input:
        print("ERROR: --input is required.", file=sys.stderr)
        return 2

    in_path = Path(args.input)
    if not in_path.exists():
        print(f"ERROR: input not found: {in_path}", file=sys.stderr)
        return 2

    out_path = Path(args.output) if args.output else in_path.with_suffix(".txt")

    # ---- Load source & lookups ----
    hs_units = engine.load_hs_units(args.cd_blank)
    iso2_to_name = {iso2: full for full, iso2 in engine.COUNTRY_NAME_TO_ISO2.items()}

    wb = openpyxl.load_workbook(in_path, data_only=True)
    meta = _scan_inv_metadata(wb)

    data_sheet = _pick_data_sheet(wb, args.hssc_sheet)
    ws = wb[data_sheet]

    # Aggregate: 'HS CODE SUM' / 'HSSC' are pre-aggregated; otherwise use INV.
    if data_sheet in ("HS CODE SUM", "HSSC"):
        agg_rows = engine.aggregate_hs_code_sum(ws, iso2_to_name)
    else:
        agg_rows = engine.aggregate_inv(ws)

    if not agg_rows:
        print(f"No data rows found in '{data_sheet}'. This looks like a BLANK "
              f"template (all HS=0 / country=#N/A).\nProvide a filled "
              f"{profile.company_name} workbook with real HS CODE SUM / HSSC data.",
              file=sys.stderr)
        return 1

    # ---- Resolve IH header values (priority: CLI > source > profile) ----
    decl_no = args.decl_no or meta.get("decl_no") or in_path.stem
    decl_date = args.decl_date or _fmt_date(meta.get("date")) \
        or _dt.date.today().strftime("%Y-%m-%d")
    company = args.company_name_override or profile.company_name
    supplier_code = args.supplier_code if args.supplier_code is not None \
        else profile.supplier_code
    incoterm = args.incoterm or profile.incoterm_default
    total = args.total if args.total is not None \
        else sum(r["amount"] for r in agg_rows)

    if args.pages is not None:
        pages = args.pages
    else:
        # Heuristic from INV row count; warn that it's approximate.
        try:
            pages = engine.count_inv_pages(wb)
        except Exception:
            pages = 1

    # ---- Currency note: engine writes profile currency into IH[8] ----
    # generate_txt currently hard-codes 'USD'; inject the profile currency.
    txt = _generate_with_currency(
        agg_rows,
        decl_no=str(decl_no),
        decl_date=str(decl_date),
        pages=pages,
        company=company,
        supplier_code=supplier_code,
        total_usd=total,
        incoterm=incoterm,
        hs_units=hs_units,
        currency=profile.currency,
    )

    out_path.write_bytes(txt.encode("utf-8"))

    # ---- Optionally fill a template copy ----
    if args.fill_wens:
        import shutil
        shutil.copy(in_path, args.fill_wens)
        wb2 = openpyxl.load_workbook(args.fill_wens)
        engine.write_hssc_sheet(wb2, agg_rows)
        engine.write_piv_hsc_sheet(wb2, agg_rows)
        wb2.save(args.fill_wens)
        wb2.close()

    wb.close()

    # ---- Report ----
    print(f"Company:    {profile.company_name}  [{profile.key}]"
          f"{'' if profile.calibrated else '  (⚠️ uncalibrated)'}")
    print(f"Source:     {in_path.name}  (sheet: {data_sheet})")
    print(f"Output:     {out_path}")
    print(f"Rows:       {len(agg_rows)} (HS×Country pairs)")
    print(f"Currency:   {profile.currency}")
    print(f"Decl no:    {decl_no}")
    print(f"Date:       {decl_date}")
    print(f"Pages:      {pages}")
    print(f"Supplier #: {supplier_code}")
    print(f"Total:      {total:.2f}")

    missing = sorted({r["hs"] for r in agg_rows
                      if engine.normalize_hs(r["hs"], hs_units) not in hs_units})
    if missing:
        print(f"\n⚠️  HS codes not in dt_hscodes (treated as kg): {missing}")
    return 0


def _generate_with_currency(agg_rows, *, currency, **kw) -> str:
    """Wrapper that runs engine.generate_txt then swaps the IH currency field
    if the profile uses something other than USD.

    (Kept as a thin shim so the calibrated engine stays untouched.)
    """
    txt = engine.generate_txt(agg_rows, **kw)
    if currency != "USD":
        lines = txt.split("\n")
        if lines and lines[0].startswith('"IH"'):
            parts = lines[0].split(",")
            # IH currency is field index 8
            if len(parts) > 8 and parts[8] == '"USD"':
                parts[8] = f'"{currency}"'
                lines[0] = ",".join(parts)
        txt = "\n".join(lines)
    return txt


if __name__ == "__main__":
    raise SystemExit(main())
