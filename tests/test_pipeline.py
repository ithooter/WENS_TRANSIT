"""Смоук-тесты единого конвейера: адаптер → canonical → заполнение/извлечение/SI."""
import openpyxl
import pytest

from app.engines.sources import read_shipment
from app.engines.template_select import select_template, OPTIAUTO_WHOLE
from app.engines.fillers import fill_optiauto_whole
from app.engines.si_bridge import build_si_inputs


# --------------------------- адаптер / canonical --------------------------- #

def test_combined_reads_into_canonical(combined_small):
    sh = read_shipment([combined_small])
    assert sh.source_format == "combined"
    assert sh.total_rows == 5
    assert sh.total_places == 3
    assert sh.requisites.currency == "AED"        # определилось по "Price AED"
    # перевод не выдумываем: нечётные позиции без RU остаются пустыми
    assert sh.goods[0].description_ru is None
    assert sh.goods[1].description_ru == "Товар 2"


def test_renamed_sheets_resolved_by_columns(combined_renamed):
    """Листы 'Processed Data' / 'XYZ Random' опознаются по сигнатуре колонок."""
    sh = read_shipment([combined_renamed])
    assert sh.total_rows == 5
    assert sh.total_places == 3


# ------------------------------ заполнение --------------------------------- #

def test_fill_and_extract_sheets(combined_small, tmp_path):
    sh = read_shipment([combined_small])
    whole = tmp_path / "whole.xlsx"
    rep = fill_optiauto_whole(sh, OPTIAUTO_WHOLE, whole,
                              invoice_no="T1", date="01.07.2026", remove_logo=False)
    assert rep["inv_rows"] == 5 and rep["pl_rows"] == 3

    wb = openpyxl.load_workbook(whole)
    assert wb.sheetnames == ["INV", "PL"]
    inv = wb["INV"]
    assert inv["B23"].value == 1                  # renumber N
    assert inv["C23"].value == "BRAND1"           # brand
    assert inv["N15"].value == "T1"               # invoice no в шапке
    assert inv["Q18"].value == 3                  # Num Pckg's = число коробок
    assert inv["G186"].value == "=SUBTOTAL(9,G23:G185)"   # итоги не сдвинуты (влезло)
    wb.close()


def test_expansion_keeps_totals_consistent(combined_large, tmp_path):
    """200 строк > вместимости 163 → строки вставлены, итоги пересчитаны на новый диапазон."""
    sh = read_shipment([combined_large])
    whole = tmp_path / "whole_big.xlsx"
    rep = fill_optiauto_whole(sh, OPTIAUTO_WHOLE, whole, invoice_no="T2", date=None)
    assert rep["inv_rows"] == 200 and rep["pl_rows"] == 30

    wb = openpyxl.load_workbook(whole)
    inv = wb["INV"]
    extra = 200 - 163                               # 37
    t1 = 186 + extra                                # 223
    data_end = 23 + 200 - 1                          # 222
    assert inv.cell(row=t1, column=7).value == f"=SUBTOTAL(9,G23:G{data_end})"
    assert inv.cell(row=16, column=14).value == f"=G{t1}"   # header ref обновлён
    # клонированная VLOOKUP-формула на вставленной строке сдвинута корректно
    assert inv.cell(row=100, column=20).value.startswith("=VLOOKUP(S100")
    pl = wb["PL"]
    pl_extra = 30 - 21                              # 9
    assert pl.cell(row=43 + pl_extra, column=3).value == f"=SUM(F21:F{21 + 30 - 1})"
    wb.close()


# --------------------------------- SI/Transit ------------------------------ #

def test_si_bridge_and_transit(combined_small, tmp_path):
    import generate_si_documents as si
    sh = read_shipment([combined_small])
    sh.requisites.receiver_lines = ["LLC Test", "Moscow"]
    inv_df, pl_df, details, profile = build_si_inputs(sh)
    assert len(inv_df) == 5 and len(pl_df) == 3
    assert profile in si.PROFILES

    enriched = si.enrich(inv_df)
    agg = si.aggregate(enriched)
    agg = si.distribute_places(agg, len(pl_df), enriched)
    assert int(agg["places"].sum()) == 3           # сумма мест = число коробок

    out = tmp_path / "transit.xlsx"
    si.build_invoice_transit(agg, details, profile, out,
                             blank=False, display_invoice_no="T1")
    assert out.exists() and out.stat().st_size > 0


# ------------------------------ выбор шаблона ------------------------------ #

@pytest.mark.parametrize("sender_key,expect_key,expect_remove_logo", [
    ("optiauto", "optiauto", False),
    ("carbotec", "basic", True),
    (None, "basic", True),
    ("wens", "wens", True),
])
def test_template_selection(sender_key, expect_key, expect_remove_logo):
    ch = select_template(sender_key)
    assert ch.key == expect_key
    assert ch.remove_logo == expect_remove_logo


# ------------------- регрессии дизайна (под эталон) ------------------------ #

def test_pl_receiver_row13_and_headers_preserved(combined_small, tmp_path):
    """Получатель PL — с строки 13; заголовки колонок B19/G19 не затираются."""
    sh = read_shipment([combined_small])
    sh.requisites.receiver_lines = [
        "LLC «AVR»", "Legal address: Kakhovka Street",
        "Bld 11/1 Moscow. 117303", "Russia INN 7728733327 / KPP 772701001"]
    whole = tmp_path / "w.xlsx"
    fill_optiauto_whole(sh, OPTIAUTO_WHOLE, whole, invoice_no="SI0354", date="26.06.2026")
    wb = openpyxl.load_workbook(whole)
    pl = wb["PL"]
    assert pl["B13"].value == "LLC «AVR»"                        # имя в строке 13
    assert pl["B16"].value == "Russia INN 7728733327 / KPP 772701001"  # 4-я строка
    assert pl["B19"].value == "Номер Коробки"                    # заголовки уцелели
    assert pl["G19"].value == "Вес"
    wb.close()


def test_inv_currency_header_follows_currency(combined_small, tmp_path):
    sh = read_shipment([combined_small])                          # AED (Price AED)
    p = tmp_path / "aed.xlsx"
    fill_optiauto_whole(sh, OPTIAUTO_WHOLE, p, invoice_no="X", date=None)
    assert openpyxl.load_workbook(p)["INV"]["L22"].value == " AMT AED"
    sh.requisites.currency = "USD"
    p2 = tmp_path / "usd.xlsx"
    fill_optiauto_whole(sh, OPTIAUTO_WHOLE, p2, invoice_no="X", date=None)
    assert openpyxl.load_workbook(p2)["INV"]["L22"].value == " AMT USD"


def test_optiauto_keeps_logo_and_address_images(combined_small, tmp_path):
    """OptiAuto (лого сохраняется): на INV и PL по 2 картинки (логотип + адрес)."""
    sh = read_shipment([combined_small])
    p = tmp_path / "w.xlsx"
    fill_optiauto_whole(sh, OPTIAUTO_WHOLE, p, invoice_no="X", date=None, remove_logo=False)
    wb = openpyxl.load_workbook(p)
    assert len(wb["INV"]._images) == 2
    assert len(wb["PL"]._images) == 2
    wb.close()


def test_basic_has_no_logo_images(combined_small, tmp_path):
    sh = read_shipment([combined_small])
    p = tmp_path / "w.xlsx"
    fill_optiauto_whole(sh, OPTIAUTO_WHOLE, p, invoice_no="X", date=None, remove_logo=True)
    assert len(openpyxl.load_workbook(p)["INV"]._images) == 0


def test_inv_paper_view(combined_small, tmp_path):
    """INV: сетка спрятана, freeze по строку-заголовок таблицы, рамка (print_area)."""
    sh = read_shipment([combined_small])
    p = tmp_path / "w.xlsx"
    fill_optiauto_whole(sh, OPTIAUTO_WHOLE, p, invoice_no="X", date=None)
    wb = openpyxl.load_workbook(p)
    inv = wb["INV"]
    assert inv.sheet_view.showGridLines is False
    assert inv.freeze_panes == "A23"
    assert inv.sheet_view.view == "pageBreakPreview"        # синяя рамка видна
    assert inv.print_area == "'INV'!$A$1:$T$188"            # 5 строк → ИТОГО 185+3
    assert wb["PL"].print_area == "'PL'!$A$1:$I$45"         # 3 коробки → ИТОГО 41+4


def test_parties_data_fixed():
    from app.service import parties_for_ui
    ui = parties_for_ui()
    opti = next(s for s in ui["senders"] if s["key"] == "optiauto")["block"]
    assert "]" not in opti                                        # нет лишней ]
    assert opti.count("OPTIAUTO FZE") == 1 and "OptiAuto FZE" not in opti  # нет дубля
    avr = next(r for r in ui["receivers"] if r["key"] == "LLC AVR")["block"]
    assert "«AVR»" in avr
    assert "117303Russia" not in avr                             # запятая на месте
    assert "Russia INN 7728733327 / KPP 772701001" in avr
